# -*- coding: utf-8 -*-

# import matplotlib.pyplot as plt
# import tensorflow as tf
import errno
import os
import re
import xml.etree.ElementTree as ET

import openpyxl as xl  # this is used to parse the .xlsx files
# Here I use pillow as a maintained fork of PIL
from PIL import Image

'''
    This function is a print function used to print temporary results
    
    :arg
'''
def testprint(display_data, end_cus = '\n', no_output = 0):
    if is_test:
        if not no_output:
            print(display_data, end=end_cus)
'''
    This function is used to get all the files' and sub_folders' name for further operation
    
    :args
        file_path = the path you want to explore
        is_test = whether it's a test mode or real implementation
    :returns
        image_names = list contains all the image names 
        image_paths = list contains all the images' path
        image_filename_Includingfolder = dict contains image imageName as key, [former directory, whole path] as value
        xml_names = list of xml file names 
        xml_paths = list of xml file paths
        xml_filename_Includingfolder = dict contains xml fileName as key, [former directory, whole path, #2former directory, #2whole path, #3former directory, #3whole path] as value
'''
def get_all_image_and_xml_files(file_path, is_test = 0):
    image_names = []
    image_paths = []
    image_filename_Includingfolder = {}
    xml_names = []
    xml_paths = []
    xml_filename_Includingfolder = {}
    count1 = 0
    count2 = 0
    for root, dirs, files in os.walk(file_path):
        # print(root, "consumes", end=" ")
        # print(sum(getsize(join(root, name)) for name in files), end=" ")
        # print("bytes in", len(files), "non-directory files", end=" ")
        # print(os.path.isdir(os.path.join(root, files)))
        # print(root)
        # the file_path = "./康师傅/", so each time get into another directory, should record the sub path
        if dirs: # print the dirs only when it's a real directory(not a blank directory as: [])
            testprint(dirs) # this is the directory names
            testprint(root) # :) this is the root, so don't actually need to record the path anyway
        for filesingle in files:
            filename, suffix = os.path.splitext(filesingle)
            if suffix == '.xml':
            # if 0:
                #use regular expression to get the includeing directory
                folder = re.search(r'(../康师傅/)(.*(?=\\|/))(\\*|/)(.*)', root)
                #group(2) is the first sub folder name
                #group(1) is the file_path(inputed)
                #group(0) is all combined
                #group(4) is the second sub folder name
                # test codes
                # testprint(folder.group(0))
                # testprint(folder.group(1))
                # testprint(folder.group(2))
                # testprint(folder.group(3))
                # testprint(folder.group(4))
                if folder: #it may be None if not match in RE
                    #if the .xml is not the only one, actually, they are not:
                    if (str(filename + suffix)) in xml_filename_Includingfolder:
                        # testprint("gocha!")
                        if folder.group(4):
                            xml_filename_Includingfolder[str(filename + suffix)].extend([str(folder.group(4)), str( file_path + folder.group(2) + '/' + folder.group(4) + '/' )])
                            xml_names.append(str(filename+suffix))
                            xml_paths.append(str( file_path + folder.group(2) + '/' + folder.group(4) + '/' ))
                        elif folder.group(2):
                            xml_filename_Includingfolder[str(filename + suffix)].extend([str(folder.group(2)), str( file_path + folder.group(2) )])
                            xml_names.append(str(filename+suffix))
                            xml_paths.append(str( file_path + folder.group(2) ))
                        else:
                            xml_filename_Includingfolder[str(filename + suffix)].extend([str(file_path), str(file_path)])
                            xml_names.append(str(filename+suffix))
                            xml_paths.append(str(file_path))
                    else:
                        if folder.group(4):
                            xml_filename_Includingfolder[str(filename + suffix)] = [str(folder.group(4)), str(file_path + folder.group(2) + '/' + folder.group(4) + '/' )]
                            xml_names.append(str(filename+suffix))
                            xml_paths.append(str(file_path + folder.group(2) + '/' + folder.group(4) + '/' ))
                        elif folder.group(2):
                            xml_filename_Includingfolder[str(filename + suffix)] = [str(folder.group(2)), str(file_path + folder.group(2) )]
                            xml_names.append(str(filename+suffix))
                            xml_paths.append(str(file_path + folder.group(2) ))
                        else:
                            xml_filename_Includingfolder[str(filename + suffix)] = [str(file_path), str(file_path)]
                            xml_names.append(str(filename+suffix))
                            xml_paths.append(str(file_path))
            if suffix == '.png':
            # if 0:
                #use regular expression to get the includeing directory
                folder = re.search(r'(../康师傅/)(.*(?=\\))(\\*)(.*)', root)
                if folder: #it may be None if not match in RE
                    if folder.group(4):
                        image_filename_Includingfolder[str(filename + suffix)] = [str(folder.group(4)), str( file_path + folder.group(2) + '/' + folder.group(4) + '/' )]
                        image_names.append(str(filename+suffix))
                        image_paths.append(str( file_path + folder.group(2) + '/' + folder.group(4) + '/' ))
                    elif folder.group(2):
                        image_filename_Includingfolder[str(filename + suffix)] = [str(folder.group(2)), str( file_path + folder.group(2) )]
                        image_names.append(str(filename+suffix))
                        image_paths.append(str(file_path + folder.group(2)))
                    else:
                        image_filename_Includingfolder[str(filename + suffix)] = [str(file_path), str(file_path)]
                        image_names.append(str(filename+suffix))
                        image_paths.append(str(file_path))
        # print(files)
        if '不一致图片' in dirs:
            # dirs.remove('不一致图片')  # don't visit 不一致图片 directories
            pass # you should deal with this situation
    testprint('Image Total')
    return image_names, image_paths, image_filename_Includingfolder, xml_names, xml_paths, xml_filename_Includingfolder
def get_all_xlsx_files(file_path, is_test = 0):
    xlsx_names = []
    xlsx_paths = []
    xlsx_filename_Includingfolder = {}
    count = 0
    for root, dirs, files in os.walk(file_path):
        # if dirs: # print the dirs only when it's a real directory(not a blank directory as: [])
        #     testprint("Directorys:", ' ')
        #     testprint(dirs) # this is the directory names
        #     testprint("Roots:", ' ')
        #     testprint(root) # :) this is the root, so don't actually need to record the path anyway
        for filesingle in files:
            filename, suffix = os.path.splitext(filesingle)
            # if suffix == '.xlsx':
            #     print(filename)
            #     print(suffix)
            if suffix == '.xlsx':
            # if 0:
                #use regular expression to get the includeing directory
                # folder = re.search(r'(../康师傅/)(.*(?=\\))(\\*)(.*)', root) #this is wrong
                folder = re.search(r'(../康师傅)(/)(.*)', root)
                # # test codes
                # if folder:
                #     testprint(folder.group(0))
                #     testprint(folder.group(1))
                #     testprint(folder.group(2))
                #     testprint(folder.group(3))
                if folder:
                    if folder.group(3):
                        xlsx_filename_Includingfolder[str(filename + suffix)] = [str(folder.group(3)), str(file_path + folder.group(3) + '/')]
                        xlsx_names.append(str(filename+suffix))
                        xlsx_paths.append(str( file_path + folder.group(3) + '/' ))
                        count += 1
    testprint(count)
    return xlsx_names, xlsx_paths, xlsx_filename_Includingfolder
'''
    This function is used to extract the noodle_only parts out of the original images which contains a lot of useless information
    
    :arg
        image = the original image you want to deal with(only one, and should be the image type)
    :returns
        noodle_* = the No.* noodle sub_image in the original image    
'''
def split_original_image(Image_wait_to_split):
    box_1 = (280, 90, 610, 420)
    box_2 = (680, 90, 1010, 420)
    # original_image = Image.open(fileNameWithPath) # to prevent from doing open and close operators too frequently, not here, use in main
    noodle_1 = Image_wait_to_split.crop(box_1)
    noodle_2 = Image_wait_to_split.crop(box_2)
    ##Uesd to show images on screen
    # noodle_1.show()
    # noodle_2.show()
    return noodle_1, noodle_2

<<<<<<< Updated upstream
=======
def split_original_image_imporved_version2(Image_wait_to_split, type_of_png):
    type_of_png -= 1
    Type_noodle_image_coorinates = [[(280, 90, 610, 420), (680, 90, 1010, 420)],
                                [(70, 100, 400, 430), (470, 100, 800, 430), (870, 90, 1200, 420)],
                                [(260, 90, 590, 420), (665, 90, 995, 420)],
                                [(970, 95, 400, 425), (470, 100, 800, 430), (870, 90, 1200, 420)]]
    if type_of_png == 0 or type_of_png == 2:
        box_1 = Type_noodle_image_coorinates[type_of_png][0]
        box_2 = Type_noodle_image_coorinates[type_of_png][1]
        noodle_1 = Image_wait_to_split.crop(box_1)
        noodle_2 = Image_wait_to_split.crop(box_2)
        return noodle_1, noodle_2
    else:
        box_1 = Type_noodle_image_coorinates[type_of_png][0]
        box_2 = Type_noodle_image_coorinates[type_of_png][1]
        box_3 = Type_noodle_image_coorinates[type_of_png][2]
        noodle_1 = Image_wait_to_split.crop(box_1)
        noodle_2 = Image_wait_to_split.crop(box_2)
        noodle_3 = Image_wait_to_split.crop(box_3)
        return noodle_1, noodle_2, noodle_3

>>>>>>> Stashed changes
def split_original_image_improved(Image_wait_to_split, image_type):
    box_1 = (280, 90, 610, 420)
    box_2 = (680, 90, 1010, 420)
    # original_image = Image.open(fileNameWithPath) # to prevent from doing open and close operators too frequently, not here, use in main
    noodle_1 = Image_wait_to_split.crop(box_1)
    noodle_2 = Image_wait_to_split.crop(box_2)
    ##Uesd to show images on screen
    # noodle_1.show()
    # noodle_2.show()
    return noodle_1, noodle_2

'''
    This function is an improved version of function get_half_noodle()
    
    :arg
        noodle_only = the image only with noodle 
        ...* = the coordinates of the noodle, so this function can still deal with the original images, have default values
    :returns
        ...* = all the half of the noodle image, 330 * 165
'''
def get_helf_noodle_improved(noodle_only, up_left_x = 0, up_left_y = 0, down_right_x = 330, down_right_y = 330, pixel = 330):
    box_horizontal_left = (up_left_x, up_left_y, down_right_x - pixel/2, down_right_y)
    box_horizontal_right = (up_left_x + pixel/2, up_left_y, down_right_x, down_right_y)

    box_longitudinal_up = (up_left_x, up_left_y, down_right_x, down_right_y - pixel/2)
    box_longitudinal_down = (up_left_x, up_left_y + pixel/2, down_right_x, down_right_y)

    # in this part, since the round off error, the image will have one more pixel
    # the solution is simple, add a offset to them
    # 330/4 = 82.5
    # print(up_left_y + pixel/4)        ||  82.5
    # print(down_right_y - pixel/4)     ||  247.5
    # offset = 0.3 is just some figure that works, I don't exactly know why
    box_cross_center = (up_left_x, up_left_y + pixel/4, down_right_x, down_right_y - pixel/4 - 0.3)
    box_vertical_center = (up_left_x + pixel/4, up_left_y, down_right_x - pixel/4 - 0.3, down_right_y)

    part_horizontal_left = noodle_only.crop(box_horizontal_left)
    part_horizontal_right = noodle_only.crop(box_horizontal_right)
    part_longitudinal_up = noodle_only.crop(box_longitudinal_up)
    part_longitudinal_down = noodle_only.crop(box_longitudinal_down)
    part_cross_center = noodle_only.crop(box_cross_center)
    part_vertical_center = noodle_only.crop(box_vertical_center)

    return part_horizontal_left, part_horizontal_right, part_longitudinal_up, part_longitudinal_down,\
            part_cross_center, part_vertical_center

def get_half_noodle(image):
    pixel = 330
    # the first one
    box_1_horizontal_left = (280, 90, 610 - pixel/2, 420)
    box_1_horizontal_right = (280 - pixel/2, 90, 610, 420)

    box_1_longitudinal_up = (280, 90, 610, 420 - pixel/2)
    box_1_longitudinal_down = (280, 90, 610 - pixel/2, 420)

    box_1_cross_center = (280, 90 + pixel/4, 610, 420 - pixel/4)
    box_1_vertical_center = (280 + pixel/4, 90, 610 - pixel/4, 420)

    #the second one
    box_2_horizontal_left = (680, 90, 1010 - pixel/2, 420)
    box_2_horizontal_right = (680 - pixel/2, 90, 1010, 420)

    box_2_longitudinal_up = (680, 90, 1010, 420 - pixel/2)
    box_2_longitudinal_down = (680, 90, 1010 - pixel/2, 420)

    box_2_cross_center = (680, 90 + pixel/4, 1010, 420 - pixel/4)
    box_2_vertical_center = (680 + pixel/4, 90, 1010 - pixel/4, 420)

    part_1_horizontal_left = image.crop(box_1_horizontal_left)
    part_1_horizontal_right = image.crop(box_1_horizontal_right)
    part_1_longitudinal_up = image.crop(box_1_longitudinal_up)
    part_1_longitudinal_down = image.crop(box_1_longitudinal_down)
    part_1_cross_center = image.crop(box_1_cross_center)
    part_1_vertical_center = image.crop(box_1_vertical_center)

    part_2_horizontal_left = image.crop(box_2_horizontal_left)
    part_2_horizontal_right = image.crop(box_2_horizontal_right)
    part_2_longitudinal_up = image.crop(box_2_longitudinal_up)
    part_2_longitudinal_down = image.crop(box_2_longitudinal_down)
    part_2_cross_center = image.crop(box_2_cross_center)
    part_2_vertical_center = image.crop(box_2_vertical_center)
    return part_1_horizontal_left, part_1_horizontal_right, part_1_longitudinal_up, part_1_longitudinal_down,\
            part_1_cross_center, part_1_vertical_center, part_2_horizontal_left, part_2_horizontal_right,\
            part_2_longitudinal_up, part_2_longitudinal_down, part_2_cross_center, part_2_vertical_center

def parse_xml_file(xml_file_name):
    pass

'''
    This function is used to deal with more than one images situation, and base on the function split_original_image()
    
    :arg
        process_num = the num of images that you want to peocess
        imageNames = name list of images
        imagePaths = path list of images 
        image_Name_IncludingFolder = ...
        xmlName = ...
        xmlPaths = ...
        xml_Name_IncludingFolder = ...
    :returns
        None
'''
def split_images_in_batch_and_save(process_num, imageNames, imagePaths, image_Name_IncludingFolder, xmlNames, xmlPaths, xml_Name_IncludingFolder):
    imageCount = 0
    # for image, xml in zip(imageNames, xmlNames): # they are not in the same order
    for image in imageNames:
        split_Name_and_Suffix = re.search(r'(\S*(?=\.))(\.*)(\S*)', image)
        imageNameWithoutSuffix, suffix = split_Name_and_Suffix.group(1),split_Name_and_Suffix.group(3)
        # testprint(imageNameWithoutSuffix)
        # testprint(suffix)
        xml_name_to_search = imageNameWithoutSuffix + '.xml'
        if xml_name_to_search in xmlNames:
            if process_num == imageCount:
                break
            # before save images to files, need to find out the lable by parsing the .xml files:
            # parse_xml_file(xml_Name_IncludingFolder[image][1]+ )

            original_image = Image.open(image_Name_IncludingFolder[image][1]+image)
            # here the noodle_image1 hasn't been refresh! but why is that？
            # cause you idiot used a original_image parameter in the function! Be careful
            noodle_image1, noodle_image2 = split_original_image(original_image)
            # noodle_image1, noodle_image2 = split_original_image(image_Name_IncludingFolder[image][1]+image)
            # noodle_image1.show()
            part_horizontal_left, part_horizontal_right, part_longitudinal_up, part_longitudinal_down,\
            part_cross_center, part_vertical_center = get_helf_noodle_improved(noodle_image1)

            part_longitudinal_up = part_longitudinal_up.rotate(90, resample=0, expand=1)
            part_longitudinal_down = part_longitudinal_down.rotate(90, resample=0, expand=1)
            part_vertical_center = part_vertical_center.rotate(90, resample=0, expand=1)
            # save_image_to_files_standerd(part_horizontal_left, '1', 'hl_test' + str(imageCount), './test/')
            # save_image_to_files_standerd(part_horizontal_right, '1', 'hr_test' + str(imageCount), './test/')
            # save_image_to_files_standerd(part_longitudinal_up, '1', 'lu_test' + str(imageCount), './test/')
            # save_image_to_files_standerd(part_longitudinal_down, '1', 'ld_test' + str(imageCount), './test/')
            # save_image_to_files_standerd(part_cross_center, '1', 'cc_test' + str(imageCount), './test/')
            # save_image_to_files_standerd(part_vertical_center, '1', 'vc_test' + str(imageCount), './test/')
            save_image_to_files_standerd_JPEG(part_horizontal_left, '1', 'hl_test' + str(imageCount), './test/')
            save_image_to_files_standerd_JPEG(part_horizontal_right, '1', 'hr_test' + str(imageCount), './test/')
            save_image_to_files_standerd_JPEG(part_longitudinal_up, '1', 'lu_test' + str(imageCount), './test/')
            save_image_to_files_standerd_JPEG(part_longitudinal_down, '1', 'ld_test' + str(imageCount), './test/')
            save_image_to_files_standerd_JPEG(part_cross_center, '1', 'cc_test' + str(imageCount), './test/')
            save_image_to_files_standerd_JPEG(part_vertical_center, '1', 'vc_test' + str(imageCount), './test/')

            part_horizontal_left, part_horizontal_right, part_longitudinal_up, part_longitudinal_down,\
            part_cross_center, part_vertical_center = get_helf_noodle_improved(noodle_image2)

            part_longitudinal_up = part_longitudinal_up.rotate(90, resample=0, expand=1)
            part_longitudinal_down = part_longitudinal_down.rotate(90, resample=0, expand=1)
            part_vertical_center = part_vertical_center.rotate(90, resample=0, expand=1)
            # save_image_to_files_standerd(part_horizontal_left, '1', 'hl_test' + str(imageCount), './test/')
            # save_image_to_files_standerd(part_horizontal_right, '1', 'hr_test' + str(imageCount), './test/')
            # save_image_to_files_standerd(part_longitudinal_up, '1', 'lu_test' + str(imageCount), './test/')
            # save_image_to_files_standerd(part_longitudinal_down, '1', 'ld_test' + str(imageCount), './test/')
            # save_image_to_files_standerd(part_cross_center, '1', 'cc_test' + str(imageCount), './test/')
            # save_image_to_files_standerd(part_vertical_center, '1', 'vc_test' + str(imageCount), './test/')
            save_image_to_files_standerd_JPEG(part_horizontal_left, '1', 'hl_test' + str(imageCount), './test/')
            save_image_to_files_standerd_JPEG(part_horizontal_right, '1', 'hr_test' + str(imageCount), './test/')
            save_image_to_files_standerd_JPEG(part_longitudinal_up, '1', 'lu_test' + str(imageCount), './test/')
            save_image_to_files_standerd_JPEG(part_longitudinal_down, '1', 'ld_test' + str(imageCount), './test/')
            save_image_to_files_standerd_JPEG(part_cross_center, '1', 'cc_test' + str(imageCount), './test/')
            save_image_to_files_standerd_JPEG(part_vertical_center, '1', 'vc_test' + str(imageCount), './test/')

        imageCount += 1
'''
    This function simply store the image to the corresponding folders
    
    :arg
        image = image that will be stored, it's a image itself already, can be directory stored
            ///but I think maybe it's a smarter way not to do that, cause it will slow the program down I suppose
            ///maybe just send the name and read in sub-function will be more efficient? Don't know yet.
        lable = to denote the positive or negative samples
        tag = to denote the position of image where it belong to it's original Noodle_Image
    :returns
        None
'''
def save_image_to_files_standerd(image, lable, tag, location, directory = ''):
    #lable is True/False of whether there is a folk in this sub_image, 1 for True, 0 for False, But there should be 2
    # for other types, as there will be 4 more sub_images don't belong to any type
<<<<<<< Updated upstream
    # Tag is Count_Number of images, just give them an ID
=======
    # Tag is image name without suffix
    # image is a Image.open() type
    # location is where the sub image locates at noodle_image
    if directory:
        try:
            os.makedirs(directory + '/' + str(location) + '/', exist_ok=False)
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise    # Error when creating new folder/directory
            # for imageName in images:
            #     image = Image.open(imageName)
        try:
            image.save(directory + '/' + str(location) + '/' + str(tag) + '_' + str(lable) + '.png')
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise

    else:
        path = './' + lable + '/'
        try:
            os.makedirs(path + '/' + str(location) + '/', exist_ok=False)
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise    # Error when creating new folder/directory
        try:
            image.save(path + '/' + str(location) + '/' + str(tag) + '_' + str(lable) + '.png')
            # originalImage.save(path_1 + outFile[0])  #no need regenerate original images? Maybe need, cause the origin directory is a mess
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise

def save_image_to_files_standerd_version2(image, lable, tag, location, directory = ''):
    #lable is True/False of whether there is a folk in this sub_image, 1 for True, 0 for False, But there should be 2
    # for other types, as there will be 4 more sub_images don't belong to any type
    # Tag is image name without suffix
>>>>>>> Stashed changes
    # image is a Image.open() type
    # location is where the sub image locates at noodle_image
    if directory:
        try:
<<<<<<< Updated upstream
            os.makedirs(directory + '/' + str(location) + '/', exist_ok=False)
=======
            os.makedirs(directory + '/' + str(location) + '/' + str(lable) + '/', exist_ok=False)
>>>>>>> Stashed changes
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise    # Error when creating new folder/directory
            # for imageName in images:
            #     image = Image.open(imageName)
        try:
<<<<<<< Updated upstream
            image.save(directory + '/' + str(location) + '/' + str(tag) + '_' + str(lable) + '.png')
=======
            image.save(directory + '/' + str(location) + '/' + str(lable) + '/' + str(tag) + '.png')
>>>>>>> Stashed changes
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise

    else:
        path = './' + lable + '/'
        try:
            os.makedirs(path + '/' + str(location) + '/', exist_ok=False)
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise    # Error when creating new folder/directory
        try:
            image.save(path + '/' + str(location) + '/' + str(tag) + '_' + str(lable) + '.png')
            # originalImage.save(path_1 + outFile[0])  #no need regenerate original images? Maybe need, cause the origin directory is a mess
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise

<<<<<<< Updated upstream
def save_image_to_files_standerd_JPEG(image, lable, tag, location, directory = ''):
    #lable is True/False of whether there is a folk in this sub_image, 1 for True, 0 for False, But there should be 2
    # for other types, as there will be 4 more sub_images don't belong to any type
    # Tag is Count_Number of images, just give them an ID
=======
def save_image_to_files_standerd_version3(image, lable, tag, location, directory = ''):
    #lable is True/False of whether there is a folk in this sub_image, 1 for True, 0 for False, But there should be 2
    # for other types, as there will be 4 more sub_images don't belong to any type
    # Tag is image name without suffix
>>>>>>> Stashed changes
    # image is a Image.open() type
    # location is where the sub image locates at noodle_image
    if directory:
        try:
<<<<<<< Updated upstream
            os.makedirs(directory + '/' + str(location) + '/', exist_ok=False)
=======
            os.makedirs(directory + '/' + str(lable) + '/', exist_ok=False)
>>>>>>> Stashed changes
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise    # Error when creating new folder/directory
            # for imageName in images:
            #     image = Image.open(imageName)
        try:
<<<<<<< Updated upstream
            image.save(directory + '/' + str(location) + '/' + str(tag) + '_' + str(lable) + '.jpeg')
=======
            image.save(directory + '/' + str(lable) + '/' + str(tag) + '_' + str(location) + '.png')
>>>>>>> Stashed changes
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise

    else:
        path = './' + lable + '/'
        try:
<<<<<<< Updated upstream
            path = './' + lable + '/'
=======
>>>>>>> Stashed changes
            os.makedirs(path + '/' + str(location) + '/', exist_ok=False)
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise    # Error when creating new folder/directory
        try:
            image.save(path + '/' + str(location) + '/' + str(tag) + '_' + str(lable) + '.png')
            # originalImage.save(path_1 + outFile[0])  #no need regenerate original images? Maybe need, cause the origin directory is a mess
<<<<<<< Updated upstream
=======
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise

def save_image_to_files_standerd_JPEG(image, lable, tag, location, directory = ''):
    #lable is True/False of whether there is a folk in this sub_image, 1 for True, 0 for False, But there should be 2
    # for other types, as there will be 4 more sub_images don't belong to any type
    # Tag is Count_Number of images, just give them an ID
    # image is a Image.open() type
    # location is where the sub image locates at noodle_image
    if directory:
        try:
            os.makedirs(directory + '/' + str(location) + '/', exist_ok=False)
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise    # Error when creating new folder/directory
            # for imageName in images:
            #     image = Image.open(imageName)
        try:
            image.save(directory + '/' + str(location) + '/' + str(tag) + '_' + str(lable) + '.jpeg')
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise

    else:
        try:
            path = './' + lable + '/'
            os.makedirs(path + '/' + str(location) + '/', exist_ok=False)
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise    # Error when creating new folder/directory
        try:
            # originalImage.save(path_1 + outFile[0])  #no need regenerate original images? Maybe need, cause the origin directory is a mess
>>>>>>> Stashed changes
            image.save(path + '/' + str(location) + '/' + str(tag) + '_' + str(lable) + '.jpeg')
        except OSError as e:
            if e.errno != errno.EEXIST:
                raise

'''
    This function is used to save the return image from function split_original_image() into files.
    
    :arg
        originalImage = the original image unsplit
        noodle_* = the No.* image of splits
        fileSavePath = the path you want to preserve the new input images 
        count = used for batch saving task, it will name the image with count in middle name(dafult: count = 0) 
    :returns
        None
        
'''
def save_images_to_files_process(originalImage, noodle_1, noodle_2, fileSavePath = './OutputImage/', count = 0):
    count = 0
    # for infile in sys.argv[1:]:
    outFileName, outFileSuffix = os.path.splitext(inFileName_absolute)
    outFile = [] #deal with the name, cause need the original name to locate the image in directories quickly
    temp = "noodle_lelf_" + str(count) + ".jpeg"
    outFile.append(temp)
    temp = "noodle_right_" + str(count) + ".jpeg"
    outFile.append(temp)
    temp = "original_image_"+ str(count) + ".jpeg"
    outFile.append(temp)
    # if infile != outfile:
    original_middle_path = "whole_noodle/"
    noodle_only_path = "noodle_only/"

    path_1 = fileSavePath + original_middle_path
    path_2 = fileSavePath + noodle_only_path
    # path_3 = fileSavePath + noodle_only_path
    try:
        # os.makedirs(path_1, exist_ok=False)
        os.makedirs(path_2, exist_ok=False)
        # os.makedirs(path_2, exist_ok=False)
    except OSError as e:
        if e.errno != errno.EEXIST:
            raise    # Error when creating new folder/directory

    try:
        # originalImage.save(path_1 + outFile[0])  #no need regenerate original images? Maybe need, casue the original directory is a mess
        noodle_1.save(path_2 + outFile[1])
        noodle_2.save(path_2 + outFile[2])
    except OSError as e:
        if e.errno != errno.EEXIST:
            raise

'''
    This is function to determine if the .xml can be used to lable the origin image data
    :arg
        xml_file_name
        xml_Name_IncludingFolder 
        xlsx_Name_Path_Dict
    :returns
        if this can be used?
'''
# not done yet, but I don't think it should, just use any corresponding .xml file it can get, think over it later on
def Get_image_corresponding_xml(image_Name_IncludingFolder_xlsxFolder, xml_Name_IncludingFolder):
    for image, folder_list in image_Name_IncludingFolder_xlsxFolder.items():
        xlsx_folder = folder_list[2]

    xml_file_root_folder = xml_Name_IncludingFolder[xml_file_name][0]

    main_path = "../康师傅/"
    test_path = "./test/"
    xlsxNames, xlsxPaths, xlsx_Name_IncludingFolder = get_all_xlsx_files(main_path, is_test)
    print(".xlsx name: ", xlsxNames)
    print(".xlsx path: ", xlsxPaths)
    print(".xlsx name-folder: ", xlsx_Name_IncludingFolder)
    # check the folder
    def xlsx_parse():
        pass
    xlsx = xlsxNames[0]
    xlsx_path = xlsx_Name_IncludingFolder[xlsx][1]
    file = xlsx_path + xlsx
    testprint(file)

    xlsx_sheets = xl.load_workbook(file, read_only=True)# read_only mode to deal with xlsx files

    run = 1
    row_num = run + 1
    column_num = run + 1
    xlsx_first_sheet = xlsx_sheets['Sheet1']
    run = 1
    ele1 = xlsx_first_sheet.cell(row=1, column=2)
    ele2 = xlsx_first_sheet.cell(row=1, column=3)
    ele3 = xlsx_first_sheet.cell(row=1, column=4)
    firstPersonName = ele1.value
    secondPersonName = ele2.value
    thirdPersonName = ele3.value
    while(run): #this is first layer
        ele = xlsx_first_sheet.cell(row=row_num, column=column_num)
        if ele.value == 'OK':
            testprint(firstPersonName, ': ')
            testprint(ele.value)
            row_num += 1
            column_num += 0
        elif ele.value is None:
            testprint("The .xlsx file reach the end!")
            run = 0
        else:
            testprint(' First person fail.', ' || ')
            ele = xlsx_first_sheet.cell(row=row_num, column=column_num + 1)
            if ele.value == 'OK':
                testprint(secondPersonName, ': ')
                testprint('OK')
                row_num += 1
                column_num += 0
            else:
                testprint(' Second person also fail.', ' || ')
                ele = xlsx_first_sheet.cell(row=row_num, column=column_num + 2)
                if ele.value == 'OK':
                    testprint(thirdPersonName, ': ')
                    testprint('OK')
                    row_num += 1
                    column_num += 0
                else:
                    testprint(' ALL FAILED: Located in ', end_cus=' Row: ')
                    testprint(row_num, end_cus='')
                    testprint(' Column: ', end_cus='')
                    testprint(column_num)
            row_num += 1

def Relate_XML_with_Name_Path(image_Name_IncludingFolder):
        image_Name_IncludingFolder_XLSXfolder = {}
        for imageName, [last_layer_folder, whole_path] in image_Name_IncludingFolder.items():
            folder = re.search(r'(../康师傅/)(.*?(?=/|\\))(.*)', whole_path)
            # test print
            # testprint(folder.group(2))
            # the third part of the list is the folder where .xlsx in,
            image_Name_IncludingFolder_XLSXfolder[imageName] = [last_layer_folder, whole_path, folder.group(2)]
        #@@@@@
        xml_Folder_Name_wholePath = {}
        for xml, list in xml_Name_IncludingFolder.items():
            if len(list) == 2:
                [xmlfolder, xml_whole_path] = list
                folder1 = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=\d))(.*)', xml_whole_path)
                if not folder1:
                    testprint("NONONO!")
                if folder1:
                    testprint(folder1.group(2), end_cus=' ')
                    testprint(folder1.group(4))
                if folder1:
                    if folder1.group(2) not in xml_Folder_Name_wholePath:
                        # xml_Folder_Name_wholePath[folder1.group(2)] = [xml, xmlfolder, xml_whole_path]
                        xml_Folder_Name_wholePath[folder1.group(2)] = [{xml: [folder1.group(4), xml_whole_path]}]
                    else: # in fact after change list elements into dict, this situation will not occur, NOO! now, occur another!
                        # xml_Folder_Name_wholePath[folder1.group(2)].extend([xml, xmlfolder, xml_whole_path])
                        xml_Folder_Name_wholePath[folder1.group(2)].append([{xml: [folder1.group(4), xml_whole_path]}])
            elif len(list) == 4:
                [xmlfolder, xml_whole_path, xmlfolder2, xml_whole_path2] = list
                folder1 = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=\d))(.*)', xml_whole_path)
                folder2 = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=\d))(.*)', xml_whole_path2)
                if not folder1:
                    testprint("NONONO!")
                if folder1:
                    testprint(folder1.group(2), end_cus=' ')
                    testprint(folder1.group(4))
                if folder2:
                    testprint(folder2.group(2), end_cus=' ')
                    testprint(folder2.group(4))
                if folder1:
                    if folder1.group(2) not in xml_Folder_Name_wholePath:
                        # xml_Folder_Name_wholePath[folder1.group(2)] = [xml, xmlfolder, xml_whole_path, xmlfolder2, xml_whole_path2]
                        xml_Folder_Name_wholePath[folder1.group(2)] =[{xml: [folder1.group(4), xml_whole_path, folder2.group(4), xml_whole_path2]}]
                    else: # in fact after change list elements into dict, this situation will not occur, NOO! now, occur another!
                        # xml_Folder_Name_wholePath[folder1.group(2)].extend([xml, xmlfolder, xml_whole_path, xmlfolder2, xml_whole_path2])
                        xml_Folder_Name_wholePath[folder1.group(2)].append([{xml: [folder1.group(4), xml_whole_path, folder2.group(4), xml_whole_path2]}])
            elif len(list) == 6:
                [xmlfolder, xml_whole_path, xmlfolder2, xml_whole_path2, xmlfolder3, xml_whole_path3] = list
                folder1 = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=\d))(.*)', xml_whole_path)
                folder2 = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=\d))(.*)', xml_whole_path2)
                folder3 = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=\d))(.*)', xml_whole_path3)
                if not folder1:
                    testprint("NONONO!")
                if folder1:
                    testprint(folder1.group(2), end_cus=' ')
                    testprint(folder1.group(4))
                if folder2:
                    testprint(folder2.group(2), end_cus=' ')
                    testprint(folder2.group(4))
                if folder3:
                    testprint(folder3.group(2), end_cus=' ')
                    testprint(folder3.group(4))
                if folder1:
                    if folder1.group(2) not in xml_Folder_Name_wholePath:
                        # xml_Folder_Name_wholePath[folder1.group(2)] = [xml, xmlfolder, xml_whole_path, xmlfolder2, xml_whole_path2, xmlfolder3, xml_whole_path3]
                        xml_Folder_Name_wholePath[folder1.group(2)] = [{xml: [folder1.group(4), xml_whole_path, folder2.group(4), xml_whole_path2, folder3.group(4), xml_whole_path3]}]
                    else:# in fact after change list elements into dict, this situation will not occur, NOO! now, occur another!
                        # xml_Folder_Name_wholePath[folder1.group(2)].extend([xml, xmlfolder, xml_whole_path, xmlfolder2, xml_whole_path2, xmlfolder3, xml_whole_path3])
                        xml_Folder_Name_wholePath[folder1.group(2)].append([{xml: [folder1.group(4), xml_whole_path, folder2.group(4), xml_whole_path2, folder3.group(4), xml_whole_path3]}])
            # folder = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=(\d|\s)))(.*)', xml_whole_path)

        testprint("After trans, .xml dict size: ", end_cus='')
        testprint(len(xml_Folder_Name_wholePath))
        testprint(xml_Folder_Name_wholePath)

        # print("image's corresponding .xml file root folder/ last-last-layer-folder: ", image_Name_IncludingFolder_XMLfolder)
        # split_images_in_batch_and_save(len(imageName), imageNames, imagePaths, image_Name_IncludingFolder, xmlNames, xmlPaths, xml_Name_IncludingFolder)
        # Get_image_corresponding_xml(image_Name_IncludingFolder_XLSXfolder, xml_Name_IncludingFolder)
        for key, value in sorted(xml_Folder_Name_wholePath.items()):
            try:
              testCount
            except NameError:
                testCount = 0
            else:
                testCount += 1
                testprint(key, end_cus=' || ')
                testprint(value)
        return xml_Folder_Name_wholePath

def parse_xml_into_dict(xmlNames, xml_Name_IncludingFolder, mute = 1):
    png_coordinate = {}
    png_name_list = []
    for xml in xmlNames:
        png_name = re.search(r'(.+(?=\.))', xml)
        # testprint(png_name.group(1))
        png_name_list.append(png_name.group(1))
        try:
            tree = ET.parse(xml_Name_IncludingFolder[xml][1] + xml)
        except IndexError:
            raise Exception("There is a .xml file doesn't have corresponding file path")
        root = tree.getroot()
        for child in root: #most of the child subjects are <rigion>
            for ch_child in child: #ch_child is either <id> or <points>
                if ch_child.tag == 'id':
                    if ch_child.text == '0': #if the <id> lable text is 0, then it's a fork!
                        show_points = 1
                        testprint(ch_child.tag, end_cus=' ', no_output=mute)
                        testprint(ch_child.text, end_cus=' | ', no_output=mute)
                else:
                    if show_points == 1: #if the <id> lable text is 0, the show_points will be set as 1
                        testprint(ch_child.tag, end_cus=': ', no_output=mute)
                        sliced_test = re.findall(r'(\d+(?=\D))', ch_child.text)
                        # sliced_test = re.findall(r'(\d*?(?=\D))', ch_child.text)
                        # testprint(sliced_test)
                        count = 0
                        x = None
                        y = None
                        for piece in sliced_test:
                            # cut_test = piece.group(2)
                            # testprint(cut_test, end_cus=' | \n')
                            # print(type(cut_test))
                            if (count % 2) == 0: #######????
                                # testprint(count)
                                x = piece
                                count += 1
                            else:
                                y = piece
                                ####better not to activate, too slow
                                testprint("<", end_cus='', no_output=mute)
                                testprint(x, end_cus='', no_output=mute)
                                testprint(",", end_cus='', no_output=mute)
                                testprint(y, end_cus='', no_output=mute)
                                testprint(")", end_cus='--', no_output=mute)
                                ####better not to activate, too slow
                                # print("(", x, ",", y, ")", end='--')
                                if png_name.group(1) not in png_coordinate:
                                    png_coordinate[png_name.group(1)] = [x, y]
                                else:
                                    png_coordinate[png_name.group(1)].append([x, y])
                                count += 1
                        show_points = 0
        #open the corresponding .png file to process
    testprint(png_coordinate)
    testprint(png_name_list)
    return png_coordinate, png_name_list

<<<<<<< Updated upstream
#@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@@ to be continued
=======
#This version solve the problem that the "不一致图片" folders constaining .xml files with same file names, it will cause the coordinates wrong
>>>>>>> Stashed changes
def parse_xml_into_dict_version2(xmlNames, xml_Name_IncludingFolder, mute = 1):
    png_coordinate = {}
    png_name_list = []
    for xml in xmlNames:
        png_name = re.search(r'(.+(?=\.))', xml)
        # testprint(png_name.group(1))
        png_name_list.append(png_name.group(1))
<<<<<<< Updated upstream
        try:
            tree = ET.parse(xml_Name_IncludingFolder[xml][1] + xml)
        except IndexError:
            raise Exception("There is a .xml file doesn't have corresponding file path")
        root = tree.getroot()
        for child in root: #most of the child subjects are <rigion>
            for ch_child in child: #ch_child is either <id> or <points>
                sub_image_location_count = -1
                sub_image_location_count_last_time = 0
                if ch_child.tag == 'id':
                    sub_image_location_count += 1
                    if ch_child.text == '0': #if the <id> lable text is 0, then it's a fork!
=======
        # testprint("Now I am in file:", end_cus=' ')
        # testprint(png_name.group(1), end_cus='-->')
        which_path = 1
        if xml_Name_IncludingFolder[xml][which_path-1] == '不一致图片':
            which_path = 3
        try:
            tree = ET.parse(xml_Name_IncludingFolder[xml][which_path] + xml)
        except IndexError:
            raise Exception("There is a .xml file doesn't have corresponding file path")
        root = tree.getroot()
        # testprint("Initialize sub_image_location_count !")
        sub_image_location_count = -1
        sub_image_location_count_last_time = 0
        for child in root: #most of the child subjects are <rigion>
            # testprint("I Found a child", end_cus="-->")
            for ch_child in child: #ch_child is either <id> or <points>
                # testprint("I Found a ch_child", end_cus="-->")
                if ch_child.tag == 'id':
                    if ch_child.text == '0': #if the <id> lable text is 0, then it's a fork!
                        # testprint("I Found A Folk, sub_image_location_count ++", end_cus="-->")
                        sub_image_location_count += 1
>>>>>>> Stashed changes
                        show_points = 1 # if the <id> lable text is 0, then get coordinates following the very next
                        testprint(ch_child.tag, end_cus=' ', no_output=mute)
                        testprint(ch_child.text, end_cus=' | ', no_output=mute)
                else:
                    if show_points == 1: #if the <id> lable text is 0, the show_points will be set as 1
                        testprint(ch_child.tag, end_cus=': ', no_output=mute)
                        sliced_test = re.findall(r'(\d+(?=\D))', ch_child.text)
                        # sliced_test = re.findall(r'(\d*?(?=\D))', ch_child.text)
                        # testprint(sliced_test)
                        count = 0
                        x = None
                        y = None
                        for piece in sliced_test:
                            # cut_test = piece.group(2)
                            # testprint(cut_test, end_cus=' | \n')
                            # print(type(cut_test))
                            if (count % 2) == 0: #######????
                                # testprint(count)
                                x = piece
                                count += 1
                            else:
                                y = piece
                                ####better not to activate, too slow
                                testprint("<", end_cus='', no_output=mute)
                                testprint(x, end_cus='', no_output=mute)
                                testprint(",", end_cus='', no_output=mute)
                                testprint(y, end_cus='', no_output=mute)
                                testprint(")", end_cus='--', no_output=mute)
                                ####better not to activate, too slow
                                # print("(", x, ",", y, ")", end='--')
                                if png_name.group(1) not in png_coordinate:
                                    png_coordinate[png_name.group(1)] = [[x, y]]
                                else:
<<<<<<< Updated upstream
                                    if sub_image_location_count_last_time == sub_image_location_count:
                                        png_coordinate[png_name.group(1)][sub_image_location_count].append([x, y])
                                    else:
                                        png_coordinate[png_name.group(1)].extend([x, y])
                                count += 1
                            sub_image_location_count_last_time = sub_image_location_count
                        show_points = 0
        #open the corresponding .png file to process
    testprint(png_coordinate)
    testprint(png_name_list)
=======
                                    png_coordinate[png_name.group(1)].append([x, y])
                                    # if sub_image_location_count_last_time == sub_image_location_count:
                                    #     png_coordinate[png_name.group(1)].append([x, y])
                                    # else:
                                    #     png_coordinate[png_name.group(1)][sub_image_location_count].extend([x, y])
                                count += 1
                        sub_image_location_count_last_time = sub_image_location_count
                        # testprint("sub_image_location_count_last_time = sub_image_location_count", end_cus="-->")
                        show_points = 0
        #open the corresponding .png file to process
    # for item in png_coordinate[png_name.group(1)]:
    #     testprint(png_coordinate)
    testprint(png_coordinate, no_output=1)
    testprint(png_name_list, no_output=1)
>>>>>>> Stashed changes
    return png_coordinate, png_name_list

if __name__ == '__main__':

#############Switch###############
    process = 'prototype6'
##################################


    ###prototype 1 deal with one image:
    if (process == 'prototype1'):
        origin_iamge_path =  "../康师傅/2017.9.21/2017.9.21原图/"
        inFileName_relative = "leaper_09_20170114_130001131_0000005624.png"
        inFileName_absolute = origin_iamge_path + inFileName_relative
        original_image = Image.open(inFileName_absolute)#Image.open(origin_iamge_path + "leaper_09_20170114_130001131_0000005624.png")
        print(original_image.format, original_image.size, original_image.mode)
        noodle_1_image, noodle_2_image = split_original_image(original_image)
        # show images: origin, split left, split right
        # original_image.show()
        # noodle_1_image.show()
        # noodle_2_image.show()
        # save_images_to_files_process(original_image, noodle_1_image, noodle_2_image)
        i1, i2, i3, i4, i5, i6 = get_helf_noodle_improved(noodle_1_image)
        # i1.show()
        # i2.show()
        # i3.show()
        # i4.show()
        i5.show()
        i6.show()
        print("i1_image_size: ",i1.size, end=' || ')
        print("i2_image_size: ",i2.size, end=' || ')
        print("i3_image_size: ",i3.size, end=' || ')
        print("i4_image_size: ",i4.size, end=' || ')
        print("i5_image_size: ",i5.size, end=' || ')
        print("i6_image_size: ",i6.size)


    ###prototype 2 image_batch_process:
    if (process == 'prototype2'):
        is_test = 1 #this is used for marking if it's in a test or not
        main_path = "../康师傅/"
        test_path = "./test/"
        imageNames, imagePaths, image_Name_IncludingFolder, xmlNames, xmlPaths, xml_Name_IncludingFolder = get_all_image_and_xml_files(main_path, is_test)
        # print('xml name: ', xmlNames)
        # print('xml path: ', xmlPaths)
        # print('xml name-folder: ', xml_Name_IncludingFolder)
        print("image name: ", imageNames)
        print("image path: ", imagePaths)
        print("image name-folder: ", image_Name_IncludingFolder)
        # testprint("This is a test: ")
        # testprint(image_Name_IncludingFolder[imageNames[0]][1])
        imageCount = len(imageNames)
        if len(imageNames) == len(imagePaths):
            print("There are ", imageCount, " images to be processed")
        # split_images_in_batch_and_save(len(imageNames), imageNames, imagePaths, image_Name_IncludingFolder, xmlNames, xmlPaths, xml_Name_IncludingFolder)
        split_images_in_batch_and_save(5, imageNames, imagePaths, image_Name_IncludingFolder, xmlNames, xmlPaths, xml_Name_IncludingFolder)

    ###prototype 3 .xlsx .xml parsing:
    if (process == 'prototype3'):
        is_test = 1 #this is used for marking if it's in a test or not
        main_path = "../康师傅/"
        test_path = "./test/"
        xlsxNames, xlsxPaths, xlsx_Name_IncludingFolder = get_all_xlsx_files(main_path, is_test)
        print(".xlsx name: ", xlsxNames)
        print(".xlsx path: ", xlsxPaths)
        print(".xlsx name-folder: ", xlsx_Name_IncludingFolder)
        # check the folder
        def xlsx_parse():
            pass
        xlsx = xlsxNames[0]
        xlsx_path = xlsx_Name_IncludingFolder[xlsx][1]
        file = xlsx_path + xlsx
        testprint(file)
        xlsx_sheets = xl.load_workbook(file, read_only= True) # read_only mode to deal with xlsx files

        run = 1
        row_num = run + 1
        column_num = run +1
        xlsx_first_sheet = xlsx_sheets['Sheet1']
        # replace all the testprint(ele.value) into detective result
        ele1 = xlsx_first_sheet.cell(row=1, column=2)
        ele2 = xlsx_first_sheet.cell(row=1, column=3)
        ele3 = xlsx_first_sheet.cell(row=1, column=4)
        firstPersonName = ele1.value
        secondPersonName = ele2.value
        thirdPersonName = ele3.value
        while(run): #this is first layer
            ele = xlsx_first_sheet.cell(row=row_num, column=column_num)
            if ele.value == 'OK':
                testprint(firstPersonName, ': ')
                testprint(ele.value)
                row_num += 1
                column_num += 0
            elif ele.value is None:
                testprint("The .xlsx file reach the end!")
                run = 0
            else:
                testprint(' First person fail.', ' || ')
                ele = xlsx_first_sheet.cell(row=row_num, column=column_num + 1)
                if ele.value == 'OK':
                    testprint(secondPersonName, ': ')
                    testprint('OK')
                    row_num += 1
                    column_num += 0
                else:
                    testprint(' Second person also fail.', ' || ')
                    ele = xlsx_first_sheet.cell(row=row_num, column=column_num + 2)
                    if ele.value == 'OK':
                        testprint(thirdPersonName, ': ')
                        testprint('OK')
                        row_num += 1
                        column_num += 0
                    else:
                        testprint(' ALL FAILED: Located in ', end_cus=' Row: ')
                        testprint(row_num, end_cus= '')
                        testprint(' Column: ', end_cus='')
                        testprint(column_num)
                row_num += 1

    ###prototype 4 .xml parsing_part1
    if (process == 'prototype4'):
        is_test = 0 #this is used for marking if it's in a test or not
        main_path = "../康师傅/"
        test_path = "./test/"
        imageNames, imagePaths, image_Name_IncludingFolder, xmlNames, xmlPaths, xml_Name_IncludingFolder = get_all_image_and_xml_files(main_path, is_test)
        print('xml name: ', xmlNames)
        print('xml path: ', xmlPaths)
        print('xml name-folder: ', xml_Name_IncludingFolder)
        testprint(".xml files number in dict is: ", end_cus='')
        testprint(len(image_Name_IncludingFolder))
        # testprint("This is a test: ")
        # testprint(image_Name_IncludingFolder[imageNames[0]][1])
        imageCount = len(imageNames)
        if len(imageNames) == len(imagePaths):
            print("There are ", imageCount, " .png images to be processed")
        xmlCount = len(xmlNames)
        if len(xmlNames) == len(xmlPaths):
            print("There are ", xmlCount, " .xml files found, And the correct number should be 11846!")
            print("The size of dict is:", len(xml_Name_IncludingFolder))
        print(xml_Name_IncludingFolder['leaper_09_20170114_130825709_0000006014.xml'])
        # split_images_in_batch_and_save(len(imageNames), imageNames, imagePaths, image_Name_IncludingFolder, xmlNames, xmlPaths, xml_Name_IncludingFolder)
        imageNames.sort()
        imagePaths.sort()
        xmlNames.sort()
        xmlPaths.sort()
        # print("image name: ", imageNames[:10])
        # print("image path: ", imagePaths[:10])
        # print("image name-folder: ", image_Name_IncludingFolder)
        #@@@@@
        image_Name_IncludingFolder_XLSXfolder = {}
        for imageName, [last_layer_folder, whole_path] in image_Name_IncludingFolder.items():
            folder = re.search(r'(../康师傅/)(.*?(?=/|\\))(.*)', whole_path)
            # test print
            # testprint(folder.group(2))
            # the third part of the list is the folder where .xlsx in,
            image_Name_IncludingFolder_XLSXfolder[imageName] = [last_layer_folder, whole_path, folder.group(2)]
        #@@@@@
        xml_Folder_Name_wholePath = {}
        for xml, list in xml_Name_IncludingFolder.items():
            if len(list) == 2:
                [xmlfolder, xml_whole_path] = list
                folder1 = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=\d))(.*)', xml_whole_path)
                if not folder1:
                    testprint("NONONO!")
                if folder1:
                    testprint(folder1.group(2), end_cus=' ')
                    testprint(folder1.group(4))
                if folder1:
                    if folder1.group(2) not in xml_Folder_Name_wholePath:
                        # xml_Folder_Name_wholePath[folder1.group(2)] = [xml, xmlfolder, xml_whole_path]
                        xml_Folder_Name_wholePath[folder1.group(2)] = [{xml: [folder1.group(4), xml_whole_path]}]
                    else: # in fact after change list elements into dict, this situation will not occur, NOO! now, occur another!
                        # xml_Folder_Name_wholePath[folder1.group(2)].extend([xml, xmlfolder, xml_whole_path])
                        xml_Folder_Name_wholePath[folder1.group(2)].append([{xml: [folder1.group(4), xml_whole_path]}])
            elif len(list) == 4:
                [xmlfolder, xml_whole_path, xmlfolder2, xml_whole_path2] = list
                folder1 = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=\d))(.*)', xml_whole_path)
                folder2 = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=\d))(.*)', xml_whole_path2)
                if not folder1:
                    testprint("NONONO!")
                if folder1:
                    testprint(folder1.group(2), end_cus=' ')
                    testprint(folder1.group(4))
                if folder2:
                    testprint(folder2.group(2), end_cus=' ')
                    testprint(folder2.group(4))
                if folder1:
                    if folder1.group(2) not in xml_Folder_Name_wholePath:
                        # xml_Folder_Name_wholePath[folder1.group(2)] = [xml, xmlfolder, xml_whole_path, xmlfolder2, xml_whole_path2]
                        xml_Folder_Name_wholePath[folder1.group(2)] =[{xml: [folder1.group(4), xml_whole_path, folder2.group(4), xml_whole_path2]}]
                    else: # in fact after change list elements into dict, this situation will not occur, NOO! now, occur another!
                        # xml_Folder_Name_wholePath[folder1.group(2)].extend([xml, xmlfolder, xml_whole_path, xmlfolder2, xml_whole_path2])
                        xml_Folder_Name_wholePath[folder1.group(2)].append([{xml: [folder1.group(4), xml_whole_path, folder2.group(4), xml_whole_path2]}])
            elif len(list) == 6:
                [xmlfolder, xml_whole_path, xmlfolder2, xml_whole_path2, xmlfolder3, xml_whole_path3] = list
                folder1 = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=\d))(.*)', xml_whole_path)
                folder2 = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=\d))(.*)', xml_whole_path2)
                folder3 = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=\d))(.*)', xml_whole_path3)
                if not folder1:
                    testprint("NONONO!")
                if folder1:
                    testprint(folder1.group(2), end_cus=' ')
                    testprint(folder1.group(4))
                if folder2:
                    testprint(folder2.group(2), end_cus=' ')
                    testprint(folder2.group(4))
                if folder3:
                    testprint(folder3.group(2), end_cus=' ')
                    testprint(folder3.group(4))
                if folder1:
                    if folder1.group(2) not in xml_Folder_Name_wholePath:
                        # xml_Folder_Name_wholePath[folder1.group(2)] = [xml, xmlfolder, xml_whole_path, xmlfolder2, xml_whole_path2, xmlfolder3, xml_whole_path3]
                        xml_Folder_Name_wholePath[folder1.group(2)] = [{xml: [folder1.group(4), xml_whole_path, folder2.group(4), xml_whole_path2, folder3.group(4), xml_whole_path3]}]
                    else:# in fact after change list elements into dict, this situation will not occur, NOO! now, occur another!
                        # xml_Folder_Name_wholePath[folder1.group(2)].extend([xml, xmlfolder, xml_whole_path, xmlfolder2, xml_whole_path2, xmlfolder3, xml_whole_path3])
                        xml_Folder_Name_wholePath[folder1.group(2)].append([{xml: [folder1.group(4), xml_whole_path, folder2.group(4), xml_whole_path2, folder3.group(4), xml_whole_path3]}])
            # folder = re.search(r'(../康师傅/)(.*?(?=/))(/?)(.*?(?=(\d|\s)))(.*)', xml_whole_path)

        testprint("After trans, .xml dict size: ", end_cus='')
        testprint(len(xml_Folder_Name_wholePath))
        testprint(xml_Folder_Name_wholePath)

        # print("image's corresponding .xml file root folder/ last-last-layer-folder: ", image_Name_IncludingFolder_XMLfolder)
        # split_images_in_batch_and_save(len(imageName), imageNames, imagePaths, image_Name_IncludingFolder, xmlNames, xmlPaths, xml_Name_IncludingFolder)
        # Get_image_corresponding_xml(image_Name_IncludingFolder_XLSXfolder, xml_Name_IncludingFolder)
        for key, value in sorted(xml_Folder_Name_wholePath.items()):
            try:
              testCount
            except NameError:
                testCount = 0
            else:
                testCount += 1
                print(key, end='||')
                print(value)

    ###prototype 5 .xml parsing_part2
    if (process == 'prototype5'):
        is_test = 1 #this is used for marking if it's in a test or not
        main_path = "../康师傅/"
        test_path = "./test/"

        ###Step1 begin========================================:
        imageNames, imagePaths, image_Name_IncludingFolder, xmlNames, xmlPaths, xml_Name_IncludingFolder = get_all_image_and_xml_files(main_path, is_test)
        ###Step1 end==========================================.
        print('xml name: ', xmlNames)
        print('xml path: ', xmlPaths)
        print('xml name-folder: ', xml_Name_IncludingFolder)
        testprint(".xml files number in dict is: ", end_cus='')
        testprint(len(image_Name_IncludingFolder))
        imageCount = len(imageNames)
        if len(imageNames) == len(imagePaths):
            print("There are ", imageCount, " .png images to be processed")
        xmlCount = len(xmlNames)
        if len(xmlNames) == len(xmlPaths):
            print("There are ", xmlCount, " .xml files found, And the correct number should be 11846!")
            print("The size of dict is:", len(xml_Name_IncludingFolder))
        # split_images_in_batch_and_save(len(imageNames), imageNames, imagePaths, image_Name_IncludingFolder, xmlNames, xmlPaths, xml_Name_IncludingFolder)
        ###Step2 begin========================================:
        xml_date_with_PersonName_and_path_dict = Relate_XML_with_Name_Path(image_Name_IncludingFolder)
        ###Step2 end==========================================.
        testprint(list(xml_date_with_PersonName_and_path_dict.items())[0:5])

        ######################part2
        testprint(xmlNames[0])
        testprint(xml_Name_IncludingFolder[xmlNames[0]][1])

        # for xml in xmlNames:
        #     try:
        #         xml_Name_IncludingFolder[xml][3]
        #     except:
        #         testprint(xml_Name_IncludingFolder[xml], end_cus=' || Index Error! \n')
        png_coordinate = {}
        png_name_list = []
        for xml in xmlNames:
            png_name = re.search(r'(.+(?=\.))', xml)
            # testprint(png_name.group(1))
            png_name_list.append(png_name.group(1))
            try:
                tree = ET.parse(xml_Name_IncludingFolder[xml][1] + xml)
            except IndexError:
                raise Exception("There is a .xml file doesn't have corresponding file path")
            root = tree.getroot()
            for child in root: #most of the child subjects are <rigion>
                for ch_child in child: #ch_child is either <id> or <points>
                    if ch_child.tag == 'id':
                        if ch_child.text == '0': #if the <id> lable text is 0, then it's a fork!
                            show_points = 1
                            testprint(ch_child.tag, end_cus=' ')
                            testprint(ch_child.text, end_cus=' | ')
                    else:
                        if show_points == 1: #if the <id> lable text is 0, the show_points will be set as 1
                            testprint(ch_child.tag, end_cus=': ')

                            # testprint(ch_child.text, end_cus=' | ')
                            sliced_test = re.findall(r'(\d+(?=\D))', ch_child.text)
                            # sliced_test = re.findall(r'(\d*?(?=\D))', ch_child.text)
                            # testprint(sliced_test)
                            count = 0
                            x = None
                            y = None
                            for piece in sliced_test:
                                # cut_test = piece.group(2)
                                # testprint(cut_test, end_cus=' | \n')
                                # print(type(cut_test))
                                if (count % 2) == 0:
                                    # testprint(count)
                                    x = piece
                                    count += 1
                                else:
                                    y = piece
                                    testprint("<",end_cus='')
                                    testprint(x,end_cus='')
                                    testprint(",",end_cus='')
                                    testprint(y,end_cus='')
                                    testprint(")",end_cus='--')
                                    # print("(", x, ",", y, ")", end='--')
                                    if png_name.group(1) not in png_coordinate:
                                        png_coordinate[png_name.group(1)] = [x, y]
                                    else:
                                        png_coordinate[png_name.group(1)].append([x, y])
                                    count += 1
                            show_points = 0
            #open the corresponding .png file to process
        testprint(png_coordinate)
        testprint(png_name_list)

    ###prototype 6 .image_slice
    if (process == 'prototype6'):
        is_test = 0 #this is used for marking if it's in a test or not
        main_path = "../康师傅/"
        test_path = "./test/"

        ###Step1 begin========================================:
        imageNames, imagePaths, image_Name_IncludingFolder, xmlNames, xmlPaths, xml_Name_IncludingFolder = get_all_image_and_xml_files(main_path, is_test)
        ###Step1 end==========================================.
        print('xml name: ', xmlNames)
        print('xml path: ', xmlPaths)
        print('xml name-folder: ', xml_Name_IncludingFolder)
        testprint(".xml files number in dict is: ", end_cus='')
        testprint(len(image_Name_IncludingFolder))
        imageCount = len(imageNames)
        if len(imageNames) == len(imagePaths):
            print("There are ", imageCount, " .png images to be processed")
        xmlCount = len(xmlNames)
        if len(xmlNames) == len(xmlPaths):
            print("There are ", xmlCount, " .xml files found, And the correct number should be 11846!")
            print("The size of dict is:", len(xml_Name_IncludingFolder))
        # split_images_in_batch_and_save(len(imageNames), imageNames, imagePaths, image_Name_IncludingFolder, xmlNames, xmlPaths, xml_Name_IncludingFolder)
        ###Step2 begin========================================:
        xml_date_with_PersonName_and_path_dict = Relate_XML_with_Name_Path(image_Name_IncludingFolder)
        ###Step2 end==========================================.
        testprint(list(xml_date_with_PersonName_and_path_dict.items())[0:5])

        ###Step3 begin========================================:
<<<<<<< Updated upstream
        png_coordinate, png_name_list = parse_xml_into_dict(xmlNames, xml_Name_IncludingFolder)
        ###Step3 end==========================================.
=======
        png_coordinate, png_name_list = parse_xml_into_dict_version2(xmlNames, xml_Name_IncludingFolder)
        ###Step3 end==========================================.

        # ####@T1@####to solve the .xml corresponding wrong
        # testprint("Here is a test:")
        # temp = png_name_list[0]
        # temp2 = xml_Name_IncludingFolder[png_name_list[0]+'.xml']
        # testprint(temp)
        # testprint(temp2)
        # testprint(png_coordinate[temp])
        # ####@T1@####to solve the .xml corresponding wrong

        ######################
        # image_wait_to_slice_without_suffix = []
        # for image_wait_to_slice in imageNames:
        #     image_wait_to_slice_without_suffix_one = re.search(r'(.+(?=\.))', image_wait_to_slice).group(1)
        #     image_wait_to_slice_without_suffix.append(image_wait_to_slice_without_suffix_one)
        # print("image_wait_to_slice_without_suffix:", end=' ')
        # testprint(image_wait_to_slice_without_suffix)

        # image_wait_to_slice_without_suffix has "color" inside, have to get rid of it
        delete_object = 'colors'
        # delete_count = 0
        # for items in png_name_list:
        #     if items == delete_object:
        #         delete_index = png_name_list.index(items)
        #         testprint(delete_index, end_cus=" : ")
        #         testprint(png_name_list[delete_index])
        #         png_name_list.pop(delete_index)
        #         # del png_name_list[delete_index]
        #         testprint(delete_index, end_cus=" : ")
        #         testprint(png_name_list[delete_index])
        #         delete_count += 1
        # testprint("There are ", end_cus='')
        # testprint(delete_count, end_cus=" has been deleted.\n")

        ###Step4 begin========================================:
        # png_coordinate, png_name_list
        # Type1 = [(280, 90, 610, 420), (680, 90, 1010, 420)]
        # Type2 = [(70, 100, 400, 430), (470, 100, 800, 430), (870, 90, 1200, 420)]
        # Type3 = [(260, 90, 590, 420), (665, 90, 995, 420)]
        # Type4 = [(70, 95, 400, 425), (470, 100, 800, 430), (870, 90, 1200, 420)]

        def which_type_xml_file_is(xml_file_name, xml_file_path):
            type_list = {}
            type_list['2017.9.21'] = 1
            type_list['2017.9.22'] = 1
            type_list['2017.9.25'] = 1
            type_list['2017.9.26'] = 1
            type_list['2017.9.27'] = 1
            type_list['2017.9.28'] = 1
            type_list['2017.9.29'] = 1
            type_list['2017.9.30'] = 1
            type_list['2017.10.9'] = 1
            type_list['2017.10.10'] = 1
            type_list['2017.10.11'] = 2
            type_list['2017.10.12'] = 2
            type_list['2017.10.13'] = 2
            type_list['2017.10.16'] = 2
            type_list['2017.10.17'] = 2
            type_list['2017.10.18'] = 2
            type_list['2017.10.19'] = 2
            type_list['2017.10.20'] = 2
            type_list['2017.10.23'] = 2
            type_list['2017.10.24'] = 4
            type_list['2017.10.25'] = 4
            type_list['2017.10.26'] = 4
            type_list['2017.10.27'] = 4
            type_list['2017.10.30'] = 4
            type_list['2017.10.31'] = 4
            type_list['2017.11.01'] = 4
            type_list['2017.11.02'] = 4
            type_list['2017.11.03'] = 3
            type_list['2017.11.06'] = 3
            type_list['2017.11.07'] = 3
            type_list['2017.11.09'] = 3
            type_list['2017.11.14'] = 3
            type_list['2017.11.16'] = 3
            type_list['2017.11.17'] = 1
            type_list['2017.11.20'] = 1
            type_list['2017.11.24'] = 1
            type_list['2017.12.4'] = 1
            type_list['2017.12.5'] = 1
            type_list['2017.12.6'] = 1
            type_list['2017.12.7'] = 1
            folder_search = re.search(r'(.*?(?=\d))((\d|\.)*(?=\/))(.*)', xml_file_path)
            testprint(folder_search.group(2), end_cus=" And type is: ")
            type_of_this_file = type_list[folder_search.group(2)]
            testprint(type_of_this_file)
            return folder_search.group(2), type_of_this_file

        #consider of exceptions
        def which_type_xml_file_is_improved(xml_and_png_file_name_no_Suffix, xml_file_path, excption_png_dict):
            type_list = {}
            type_list['2017.9.21'] = 1
            type_list['2017.9.22'] = 1
            type_list['2017.9.25'] = 1
            type_list['2017.9.26'] = 1
            type_list['2017.9.27'] = 1
            type_list['2017.9.28'] = 1
            type_list['2017.9.29'] = 1
            type_list['2017.9.30'] = 1
            type_list['2017.10.9'] = 1
            type_list['2017.10.10'] = 1
            type_list['2017.10.11'] = 2
            type_list['2017.10.12'] = 2
            type_list['2017.10.13'] = 2
            type_list['2017.10.16'] = 2
            type_list['2017.10.17'] = 2
            type_list['2017.10.18'] = 2
            type_list['2017.10.19'] = 2
            type_list['2017.10.20'] = 2
            type_list['2017.10.23'] = 2
            type_list['2017.10.24'] = 4
            type_list['2017.10.25'] = 4
            type_list['2017.10.26'] = 4
            type_list['2017.10.27'] = 4
            type_list['2017.10.30'] = 4
            type_list['2017.10.31'] = 4
            type_list['2017.11.01'] = 4
            type_list['2017.11.02'] = 4
            type_list['2017.11.03'] = 3
            type_list['2017.11.06'] = 3
            type_list['2017.11.07'] = 3
            type_list['2017.11.09'] = 3
            type_list['2017.11.14'] = 3
            type_list['2017.11.16'] = 3
            type_list['2017.11.17'] = 1
            type_list['2017.11.20'] = 1
            type_list['2017.11.24'] = 1
            type_list['2017.12.4'] = 1
            type_list['2017.12.5'] = 1
            type_list['2017.12.6'] = 1
            type_list['2017.12.7'] = 1
            # xml_file_name
            folder_search = re.search(r'(.*?(?=\d))((\d|\.)*(?=\/))(.*)', xml_file_path)
            testprint(folder_search.group(2), end_cus=" And type is: ")
            type_of_this_file = type_list[folder_search.group(2)]
            xml_and_png_file_name_with_Suffix = xml_and_png_file_name_no_Suffix + '.png'
            if xml_and_png_file_name_with_Suffix in excption_png_dict:
                testprint("There is an exception!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
                type_of_this_file = excption_png_dict[xml_and_png_file_name_with_Suffix]
            testprint(type_of_this_file)
            return folder_search.group(2), type_of_this_file

        def get_four_quadrant(coodinate):
            x_lu, y_lu, x_rd, y_rd = coodinate
            quadrant_1 = [(x_lu + x_rd)/2, y_lu, x_rd, (y_lu + y_rd)/2]
            quadrant_2 = [x_lu, y_lu, (x_lu + x_rd)/2, (y_lu + y_rd)/2]
            quadrant_3 = [x_lu, (y_lu + y_rd)/2, (x_lu + x_rd)/2, y_rd]
            quadrant_4 = [(x_lu + x_rd)/2, (y_lu + y_rd)/2, x_rd, y_rd]

            quadrant_1_int = []
            quadrant_2_int = []
            quadrant_3_int = []
            quadrant_4_int = []
            for ele in quadrant_1:
                quadrant_1_int.append(int(ele))
            for ele in quadrant_2:
                quadrant_2_int.append(int(ele))
            for ele in quadrant_3:
                quadrant_3_int.append(int(ele))
            for ele in quadrant_4:
                quadrant_4_int.append(int(ele))
            return quadrant_1_int, quadrant_2_int, quadrant_3_int, quadrant_4_int

        def get_six_rigion(coodinate):
            x_lu, y_lu, x_rd, y_rd = coodinate
            rigion_1 = [x_lu, y_lu, (x_lu + x_rd)/2, y_rd]
            rigion_2 = [(x_lu + x_rd)/2, y_lu, x_rd, y_rd]
            rigion_3 = [x_lu, y_lu, x_rd, (y_lu + y_rd)/2]
            rigion_4 = [x_lu, (y_lu + y_rd)/2, x_rd, y_rd]
            rigion_5 = [(3*x_lu + x_rd)/4, y_lu, (x_lu + 3*x_rd)/4, y_rd]
            rigion_6 = [x_lu, (3*y_lu + y_rd)/4, x_rd, (y_lu + 3*y_rd)/4]

            rigion_1_int = []
            rigion_2_int = []
            rigion_3_int = []
            rigion_4_int = []
            rigion_5_int = []
            rigion_6_int = []
            for ele in rigion_1:
                rigion_1_int.append(int(ele))
            for ele in rigion_2:
                rigion_2_int.append(int(ele))
            for ele in rigion_3:
                rigion_3_int.append(int(ele))
            for ele in rigion_4:
                rigion_4_int.append(int(ele))
            for ele in rigion_5:
                rigion_5_int.append(int(ele))
            for ele in rigion_6:
                rigion_6_int.append(int(ele))
            return rigion_1_int, rigion_2_int, rigion_3_int, rigion_4_int, rigion_5_int, rigion_6_int

        def get_sub_png_lable(png_image_name, sub_image_location, xml_type, corresponding_coordinate):
            #sub_image_location should be 0/1/2
            Type_noodle_image_coorinates = [[(280, 90, 610, 420), (680, 90, 1010, 420)],
                                            [(70, 100, 400, 430), (470, 100, 800, 430), (870, 90, 1200, 420)],
                                            [(260, 90, 590, 420), (665, 90, 995, 420)],
                                            [(70, 95, 400, 425), (470, 100, 800, 430), (870, 90, 1200, 420)]]
            noodle_image_coordinate = Type_noodle_image_coorinates[xml_type - 1][sub_image_location]
            coodinates = corresponding_coordinate
            quadrant_1_int, quadrant_2_int, quadrant_3_int, quadrant_4_int = get_four_quadrant(noodle_image_coordinate)
            quadrant_1_points_count, quadrant_2_points_count, quadrant_3_points_count, quadrant_4_points_count = 0
            for [x, y] in coodinates:
                counts_value = {}
                if (quadrant_1_int[0] <= x) and (quadrant_1_int[2] >= x):
                    if (quadrant_1_int[1] <= y) and (quadrant_1_int[3] >= y):
                        quadrant_1_points_count += 1
                elif (quadrant_2_int[0] <= x) and (quadrant_2_int[2] >= x):
                    if (quadrant_2_int[1] <= y) and (quadrant_2_int[3] >= y):
                        quadrant_2_points_count += 1
                elif (quadrant_3_int[0] <= x) and (quadrant_3_int[2] >= x):
                    if (quadrant_3_int[1] <= y) and (quadrant_3_int[3] >= y):
                        quadrant_3_points_count += 1
                elif (quadrant_4_int[0] <= x) and (quadrant_4_int[2] >= x):
                    if (quadrant_4_int[1] <= y) and (quadrant_4_int[3] >= y):
                        quadrant_4_points_count += 1
                counts_value['1'] = quadrant_1_points_count
                counts_value['2'] = quadrant_2_points_count
                counts_value['3'] = quadrant_3_points_count
                counts_value['4'] = quadrant_4_points_count
                if counts_value['1'] >= counts_value['2']:
                    temp1 = [counts_value['1'], 1]
                    temp2 = [counts_value['2'], 2]
                else:
                    temp1 = [counts_value['2'], 2]
                    temp2 = [counts_value['1'], 1]
                if counts_value['3'] >= counts_value['4']:
                    temp3 = [counts_value['3'], 3]
                    temp4 = [counts_value['4'], 4]
                else:
                    temp3 = [counts_value['4'], 4]
                    temp4 = [counts_value['3'], 3]
                if temp1[0] >= temp2[0]:
                    max_type = temp1
                else:
                    max_type = temp2
                if temp3[0] <= temp4[0]:
                    min_type = temp3
                else:
                    min_type = temp4
                testprint(max_type[1], end_cus=' ')
                testprint(min_type[1])
            return max_type[1], min_type[1]

        ###@@@@@@@@need to check the program
        def get_sub_png_lable_improved(sub_image_location, xml_type, corresponding_coordinate):
            #sub_image_location should be 0/1/2
            Type_noodle_image_coorinates = [[[280, 90, 610, 420], [680, 90, 1010, 420]],
                                            [[70, 100, 400, 430], [470, 100, 800, 430], [870, 90, 1200, 420]],
                                            [[260, 90, 590, 420], [665, 90, 995, 420]],
                                            [[70, 95, 400, 425], [470, 100, 800, 430], [870, 90, 1200, 420]]]

            # for noodle_image_coordinate in Type_noodle_image_coorinates[xml_type - 1]:
            noodle_image_coordinate = Type_noodle_image_coorinates[xml_type - 1][sub_image_location]
            #in every noodle part of sub_png_image
            coodinates = corresponding_coordinate
            quadrant_1_int, quadrant_2_int, quadrant_3_int, quadrant_4_int = get_four_quadrant(noodle_image_coordinate)
            quadrant_1_points_count = 0
            quadrant_2_points_count = 0
            quadrant_3_points_count = 0
            quadrant_4_points_count = 0
            for [x_str, y_str] in coodinates:###########@@@@
                x = int(x_str)
                y = int(y_str)
                counts_value = {}
                if (quadrant_1_int[0] <= x) and (quadrant_1_int[2] >= x):
                    if (quadrant_1_int[1] <= y) and (quadrant_1_int[3] >= y):
                        quadrant_1_points_count += 1
                elif (quadrant_2_int[0] <= x) and (quadrant_2_int[2] >= x):
                    if (quadrant_2_int[1] <= y) and (quadrant_2_int[3] >= y):
                        quadrant_2_points_count += 1
                elif (quadrant_3_int[0] <= x) and (quadrant_3_int[2] >= x):
                    if (quadrant_3_int[1] <= y) and (quadrant_3_int[3] >= y):
                        quadrant_3_points_count += 1
                elif (quadrant_4_int[0] <= x) and (quadrant_4_int[2] >= x):
                    if (quadrant_4_int[1] <= y) and (quadrant_4_int[3] >= y):
                        quadrant_4_points_count += 1
            testprint("The points number in each quadrant is: ", end_cus='')
            testprint(quadrant_1_points_count, end_cus="||")
            testprint(quadrant_2_points_count, end_cus="||")
            testprint(quadrant_3_points_count, end_cus="||")
            testprint(quadrant_4_points_count)
            counts_value['1'] = quadrant_1_points_count
            counts_value['2'] = quadrant_2_points_count
            counts_value['3'] = quadrant_3_points_count
            counts_value['4'] = quadrant_4_points_count
            if counts_value['1'] >= counts_value['2']:
                temp1 = [counts_value['1'], 1]
                temp2 = [counts_value['2'], 2]
            else:
                temp1 = [counts_value['2'], 2]
                temp2 = [counts_value['1'], 1]

            if counts_value['3'] >= counts_value['4']:
                temp3 = [counts_value['3'], 3]
                temp4 = [counts_value['4'], 4]
            else:
                temp3 = [counts_value['4'], 4]
                temp4 = [counts_value['3'], 3]
            if temp1[0] >= temp3[0]:
                max_type = temp1
            else:
                max_type = temp3

            if temp2[0] <= temp4[0]:
                min_type = temp2
            else:
                min_type = temp4
            # testprint(type(max_type[1]))
            # if max_type[1] > 0:
            #     testprint("Max in quadrant ", end_cus='')
            #     testprint(max_type[1], end_cus='. ')
            #     testprint("Min in quadrant ", end_cus='')
            #     testprint(min_type[1])
            #     return max_type[1], min_type[1]
            # else:
            #     testprint("The maximum value is 0!")
            if max_type[1] == 3:
                testprint("Found 3!!!!!!!!!!!!!!!!")
            return max_type[1], min_type[1]

        # this is the new version
        def get_sub_png_lable_modified(sub_image_location, xml_type, corresponding_coordinate):
            #sub_image_location should be 0/1/2
            Type_noodle_image_coorinates = [[[280, 90, 610, 420], [680, 90, 1010, 420]],
                                            [[70, 100, 400, 430], [470, 100, 800, 430], [870, 90, 1200, 420]],
                                            [[260, 90, 590, 420], [665, 90, 995, 420]],
                                            [[70, 95, 400, 425], [470, 100, 800, 430], [870, 90, 1200, 420]]]

            # for noodle_image_coordinate in Type_noodle_image_coorinates[xml_type - 1]:
            noodle_image_coordinate = Type_noodle_image_coorinates[xml_type - 1][sub_image_location]
            #in every noodle part of sub_png_image
            coodinates = corresponding_coordinate
            rigion_1_int, rigion_2_int, rigion_3_int, rigion_4_int, rigion_5_int, rigion_6_int = get_six_rigion(noodle_image_coordinate)
            rigion_1_points_count = 0
            rigion_2_points_count = 0
            rigion_3_points_count = 0
            rigion_4_points_count = 0
            rigion_5_points_count = 0
            rigion_6_points_count = 0
            miss_coordinates = 0
            for [x_str, y_str] in coodinates:###########@@@@
                x = int(x_str)
                y = int(y_str)
                counts_value = {}
                if (rigion_1_int[0] <= x) and (rigion_1_int[2] >= x):
                    if (rigion_1_int[1] <= y) and (rigion_1_int[3] >= y):
                        rigion_1_points_count += 1
                if (rigion_2_int[0] <= x) and (rigion_2_int[2] >= x):
                    if (rigion_2_int[1] <= y) and (rigion_2_int[3] >= y):
                        rigion_2_points_count += 1
                if (rigion_3_int[0] <= x) and (rigion_3_int[2] >= x):
                    if (rigion_3_int[1] <= y) and (rigion_3_int[3] >= y):
                        rigion_3_points_count += 1
                if (rigion_4_int[0] <= x) and (rigion_4_int[2] >= x):
                    if (rigion_4_int[1] <= y) and (rigion_4_int[3] >= y):
                        rigion_4_points_count += 1
                if (rigion_5_int[0] <= x) and (rigion_5_int[2] >= x):
                    if (rigion_5_int[1] <= y) and (rigion_5_int[3] >= y):
                        rigion_5_points_count += 1
                if (rigion_6_int[0] <= x) and (rigion_6_int[2] >= x):
                    if (rigion_6_int[1] <= y) and (rigion_6_int[3] >= y):
                        rigion_6_points_count += 1
                if (x < noodle_image_coordinate[0]) or (x > noodle_image_coordinate[2]) or (y < noodle_image_coordinate[1]) or (y > noodle_image_coordinate[3]):
                    miss_coordinates += 1
            testprint("The points number in each rigion are: ", end_cus='')
            testprint(rigion_1_points_count, end_cus="||")
            testprint(rigion_2_points_count, end_cus="||")
            testprint(rigion_3_points_count, end_cus="||")
            testprint(rigion_4_points_count, end_cus="||")
            testprint(rigion_5_points_count, end_cus="||")
            testprint(rigion_6_points_count)
            counts_value['1'] = rigion_1_points_count
            counts_value['2'] = rigion_2_points_count
            counts_value['3'] = rigion_3_points_count
            counts_value['4'] = rigion_4_points_count
            counts_value['5'] = rigion_5_points_count
            counts_value['6'] = rigion_6_points_count
            if counts_value['1'] >= counts_value['2']:
                temp1 = [counts_value['1'], 1]
                temp2 = [counts_value['2'], 2]
            else:
                temp1 = [counts_value['2'], 2]
                temp2 = [counts_value['1'], 1]

            if counts_value['3'] >= counts_value['4']:
                temp3 = [counts_value['3'], 3]
                temp4 = [counts_value['4'], 4]
            else:
                temp3 = [counts_value['4'], 4]
                temp4 = [counts_value['3'], 3]

            if counts_value['5'] >= counts_value['6']:
                temp5 = [counts_value['5'], 5]
                temp6 = [counts_value['6'], 6]
            else:
                temp5 = [counts_value['6'], 6]
                temp6 = [counts_value['5'], 5]

            #max in  temp1, temp3, temp5:
            if temp1[0] >= temp3[0]:
                if temp1[0] >= temp5[0]:
                    max_type = temp1
                else:
                    max_type = temp5
            else:
                if temp3[0] >= temp5[0]:
                    max_type = temp3
                else:
                    max_type = temp5

            #min in temp2, temp4, temp6
            if temp2[0] <= temp4[0]:
                if temp2[0] <= temp6[0]:
                    min_type = temp2
                else:
                    min_type = temp6
            else:
                if temp4[0] <= temp6[0]:
                    min_type = temp4
                else:
                    min_type = temp6
            # testprint(type(max_type[1]))
            # if max_type[1] > 0:
            #     testprint("Max in quadrant ", end_cus='')
            #     testprint(max_type[1], end_cus='. ')
            #     testprint("Min in quadrant ", end_cus='')
            #     testprint(min_type[1])
            #     return max_type[1], min_type[1]
            # else:
            #     testprint("The maximum value is 0!")
            if max_type[0] == 0:
                testprint("There is a no noodle sub_image!~~~~~~~~~~~~~~~~~~~~~~~~~~~~")
                return 1000, 1000
            return max_type[1], min_type[1]

        def save_png_with_lable(png_name_without_suffix, xml_type, corresponding_coordinate):
            image_Name_IncludingFolder_cannot_find_png_count = 0
            png_name_with_suffix = png_name_without_suffix + '.png'
            try:
                noodle_image_name_with_suffix_and_path = image_Name_IncludingFolder[png_name_with_suffix][1] + png_name_with_suffix
            except:
                image_Name_IncludingFolder_cannot_find_png_count += 1
            try:
                original_image = Image.open(noodle_image_name_with_suffix_and_path)
            except:
                return 0
            # noodle_image1, noodle_image2 = split_original_image_imporved(original_image, xml_type)
            if xml_type == 1 or xml_type == 3:
                noodle_image1, noodle_image2 = split_original_image_imporved_version2(original_image, xml_type)
                noodle1_label_max, noodle1_label_min = get_sub_png_lable_modified(0, xml_type, corresponding_coordinate)
                noodle2_label_max, noodle2_label_min = get_sub_png_lable_modified(1, xml_type, corresponding_coordinate)
            else:
                noodle_image1, noodle_image2, noodle_image3 = split_original_image_imporved_version2(original_image, xml_type)
                noodle1_label_max, noodle1_label_min = get_sub_png_lable_modified(0, xml_type, corresponding_coordinate)
                noodle2_label_max, noodle2_label_min = get_sub_png_lable_modified(1, xml_type, corresponding_coordinate)
                noodle3_label_max, noodle3_label_min = get_sub_png_lable_modified(2, xml_type, corresponding_coordinate)
            if xml_type == 1 or xml_type == 3:
                image_to_save = [None]*6
                image_to_save[0], image_to_save[1], image_to_save[2], image_to_save[3], image_to_save[4], image_to_save[5] = get_helf_noodle_improved(noodle_image1)
                if noodle1_label_max != 1000:
                    for i in range(6):
                        if i == int(noodle1_label_max - 1):
                            save_image_to_files_standerd_version2(image_to_save[i], 1, png_name_without_suffix, int(i + 1), '../NewData/')
                        elif i == int(noodle1_label_min - 1):
                            save_image_to_files_standerd_version2(image_to_save[i], 0, png_name_without_suffix, int(i + 1), '../NewData/')
                        else:
                            save_image_to_files_standerd_version2(image_to_save[i], 2, png_name_without_suffix, int(i + 1), '../NewData/')
                image_to_save = [None]*6
                image_to_save[0], image_to_save[1], image_to_save[2], image_to_save[3], image_to_save[4], image_to_save[5] = get_helf_noodle_improved(noodle_image2)
                if noodle2_label_max != 1000:
                    for i in range(6):
                        if i == int(noodle2_label_max - 1):
                            save_image_to_files_standerd_version2(image_to_save[i], 1, png_name_without_suffix, int(i + 1), '../NewData/')
                        elif i == int(noodle2_label_min - 1):
                            save_image_to_files_standerd_version2(image_to_save[i], 0, png_name_without_suffix, int(i + 1), '../NewData/')
                        else:
                            save_image_to_files_standerd_version2(image_to_save[i], 2, png_name_without_suffix, int(i + 1), '../NewData/')
            else:
                image_to_save = [None]*6
                image_to_save[0], image_to_save[1], image_to_save[2], image_to_save[3], image_to_save[4], image_to_save[5] = get_helf_noodle_improved(noodle_image1)
                if noodle1_label_max != 1000:
                    for i in range(6):
                        if i == int(noodle1_label_max - 1):
                            save_image_to_files_standerd_version2(image_to_save[i], 1, png_name_without_suffix, int(i + 1), '../NewData/')
                        elif i == int(noodle1_label_min - 1):
                            save_image_to_files_standerd_version2(image_to_save[i], 0, png_name_without_suffix, int(i + 1), '../NewData/')
                        else:
                            save_image_to_files_standerd_version2(image_to_save[i], 2, png_name_without_suffix, int(i + 1), '../NewData/')
                image_to_save = [None]*6
                image_to_save[0], image_to_save[1], image_to_save[2], image_to_save[3], image_to_save[4], image_to_save[5] = get_helf_noodle_improved(noodle_image2)
                if noodle2_label_max != 1000:
                    for i in range(6):
                        if i == int(noodle2_label_max - 1):
                            save_image_to_files_standerd_version2(image_to_save[i], 1, png_name_without_suffix, int(i + 1), '../NewData/')
                        elif i == int(noodle2_label_min - 1):
                            save_image_to_files_standerd_version2(image_to_save[i], 0, png_name_without_suffix, int(i + 1), '../NewData/')
                        else:
                            save_image_to_files_standerd_version2(image_to_save[i], 2, png_name_without_suffix, int(i + 1), '../NewData/')
                image_to_save = [None]*6
                image_to_save[0], image_to_save[1], image_to_save[2], image_to_save[3], image_to_save[4], image_to_save[5] = get_helf_noodle_improved(noodle_image3)
                if noodle3_label_max != 1000:
                    for i in range(6):
                        if i == int(noodle3_label_max - 1):
                            save_image_to_files_standerd_version2(image_to_save[i], 1, png_name_without_suffix, int(i + 1), '../NewData/')
                        elif i == int(noodle3_label_min - 1):
                            save_image_to_files_standerd_version2(image_to_save[i], 0, png_name_without_suffix, int(i + 1), '../NewData/')
                        else:
                            save_image_to_files_standerd_version2(image_to_save[i], 2, png_name_without_suffix, int(i + 1), '../NewData/')
            return image_Name_IncludingFolder_cannot_find_png_count

        # testprint("Here is a test:")
        # testprint(png_name_list[1])
        # testprint(png_coordinate[png_name_list[1]])
        # corresponding_coordinate_xml_path = xml_Name_IncludingFolder[png_name_list[0]+".xml"][1]
        # corresponding_coordinate_xml_folder, type_of_corresponding_coordinate_xml = which_type_xml_file_is(png_name_list[0], corresponding_coordinate_xml_path)
        # Type_noodle_image_coorinates = [[[280, 90, 610, 420], [680, 90, 1010, 420]],
        #                         [[70, 100, 400, 430], [470, 100, 800, 430], [870, 90, 1200, 420]],
        #                         [[260, 90, 590, 420], [665, 90, 995, 420]],
        #                         [[70, 95, 400, 425], [470, 100, 800, 430], [870, 90, 1200, 420]]]
        # noodle_image_coordinate_0 = Type_noodle_image_coorinates[type_of_corresponding_coordinate_xml - 1][0]
        # noodle_image_coordinate_1 = Type_noodle_image_coorinates[type_of_corresponding_coordinate_xml - 1][1]
        # a, b, c, d = get_four_quadrant(noodle_image_coordinate_0)
        # e, f, g, h = get_four_quadrant(noodle_image_coordinate_1)
        #
        # print("( ", a, ", ", b, ", ", c, ", ", d, " )")
        # print("( ", e, ", ", f, ", ", g, ", ", h, " )")

        # ####@T2@####to test the get_rigion() function ===>>> activate @T1@ included
        # testprint("Here is a test2:")
        # temp = png_name_list[0]
        # temp2 = xml_Name_IncludingFolder[png_name_list[0]+'.xml']
        # testprint(temp)
        # testprint(temp2)
        # testprint(png_coordinate[temp])
        #
        # Type_noodle_image_coorinates = [[[280, 90, 610, 420], [680, 90, 1010, 420]],
        #                 [[70, 100, 400, 430], [470, 100, 800, 430], [870, 90, 1200, 420]],
        #                 [[260, 90, 590, 420], [665, 90, 995, 420]],
        #                 [[70, 95, 400, 425], [470, 100, 800, 430], [870, 90, 1200, 420]]]
        # Q1, Q2, Q3, Q4, Q5, Q6 = get_six_rigion(Type_noodle_image_coorinates[0][0])
        # testprint("Quadrant: ", end_cus='\n')
        # testprint(Q1, end_cus='||')
        # testprint(Q2, end_cus='||')
        # testprint(Q3, end_cus='||')
        # testprint(Q4, end_cus='||')
        # testprint(Q5, end_cus='||')
        # testprint(Q6)
        # which_type_xml_file_is(temp, temp2[3])
        # get_sub_png_lable_modified(0, 1, png_coordinate[temp])
        # ####@T2@####to test the get_rigion() function ===>>> activate @T1@ as well

        #this part is used to identify the excption samples in folders
        file_path = '../pngSamples/differentTypesSamples/'
        def get_png_different_types(file_path):
            excption_png_dict = {}
            for root, dirs, files in os.walk(file_path):
                # testprint(dirs)
                for dir in dirs:
                    if dir:
                        file_path_2 = file_path + str(dir) + '/'
                        for root, dirs, files in os.walk(file_path_2):
                            for fileSingle in files:
                                type_regular = re.search(r'(.*(?=\=))(=)(\d)', dir)
                                testprint(fileSingle)
                                excption_png_dict[fileSingle] = int(type_regular.group(3))
            testprint(excption_png_dict)
            return excption_png_dict
        png_different_type_dict = get_png_different_types(file_path)

        can_not_find = 0
        corresponding_coordinate = None
        corresponding_coordinate_last_time = None
        image_Name_IncludingFolder_cannot_find_png_count = 0
        for png_image_name in png_name_list:
            if png_image_name != delete_object:
                try:
                    corresponding_coordinate = png_coordinate[png_image_name]
                    #first, judge which type the .xml file blongs to
                except:
                    #the reason that the .xml file coordinate can not be found is these .xml files are empty!!!!!!!
                    can_not_find += 1
                    # testprint("Can not find: ", end_cus='')
                    # testprint(png_image_name, end_cus='.xml\n')
                    # if (png_image_name + '.xml') in xmlNames:
                    #     testprint("But can find it in xmlNames list!\nThe origin path in xml_Name_IncludingFolder should be: ", end_cus='')
                    #     testprint(xml_Name_IncludingFolder[png_image_name + '.xml'])
                if corresponding_coordinate != corresponding_coordinate_last_time:
                    corresponding_coordinate_xml_name = png_image_name + '.xml'
                    corresponding_coordinate_xml_path = xml_Name_IncludingFolder[corresponding_coordinate_xml_name][1]
                    testprint(png_image_name, end_cus=' ')
                    corresponding_coordinate_xml_folder, type_of_corresponding_coordinate_xml = which_type_xml_file_is_improved(png_image_name, corresponding_coordinate_xml_path, png_different_type_dict)
                    # (type_of_corresponding_coordinate_xml)
                    # sub_image_location = type_of_corresponding_coordinate_xml
                    # get_sub_png_lable(png_image_name, sub_image_location, type_of_corresponding_coordinate_xml, corresponding_coordinate)
                    image_Name_IncludingFolder_cannot_find_png_count += save_png_with_lable(png_image_name, type_of_corresponding_coordinate_xml, corresponding_coordinate)
                corresponding_coordinate_last_time = corresponding_coordinate
            else:
                testprint(delete_object + " still occur")
                ##?????why after I delete the "colors", it still occur?
            # testprint(corresponding_coordinate)
        print("The count that image_Name_IncludingFolder cannot find png key: ", image_Name_IncludingFolder_cannot_find_png_count)
        #########check the error and stufff:
        # testprint("There are ", end_cus='')
        # testprint(can_not_find, end_cus=" that can not be found!\n")
        # testprint("imageNames length:")
        # testprint(len(imageNames))
        # testprint("xmlNames length:")
        # testprint(len(xmlNames))
        # testprint("xml_Name_IncludingFolder length:")
        # testprint(len(xml_Name_IncludingFolder))
        # testprint("The length of png_coordinate and png_name_list are :", end_cus='')
        # testprint(len(png_coordinate), end_cus=' ')
        # testprint(len(png_name_list))
        # png_name_list_new = set(png_name_list)
        # testprint("After set, the length of png_name_list_new is : ", end_cus='')
        # testprint(len(png_name_list))
        ###Step4 end==========================================.
        # xml_file_path
        # folder_search = re.search(r'(.*(?=\d))(\d*(?=\/)(.*))' ,xml_file_path)
        # testprint(folder_search.group(2))
        # testprint(image_Name_IncludingFolder)
        # get_png_lable(png_name_list[0])

    # Type_noodle_image_coorinates = [[(280, 90, 610, 420), (680, 90, 1010, 420)],
    #                             [(70, 100, 400, 430), (470, 100, 800, 430), (870, 90, 1200, 420)],
    #                             [(260, 90, 590, 420), (665, 90, 995, 420)],
    #                             [(970, 95, 400, 425), (470, 100, 800, 430), (870, 90, 1200, 420)]]
    #
    # print(Type_noodle_image_coorinates[0])
    # print(Type_noodle_image_coorinates[0][1])
>>>>>>> Stashed changes


        # ######################
        # # image_wait_to_slice_without_suffix = []
        # # for image_wait_to_slice in imageNames:
        # #     image_wait_to_slice_without_suffix_one = re.search(r'(.+(?=\.))', image_wait_to_slice).group(1)
        # #     image_wait_to_slice_without_suffix.append(image_wait_to_slice_without_suffix_one)
        # # print("image_wait_to_slice_without_suffix:", end=' ')
        # # testprint(image_wait_to_slice_without_suffix)
        #
        # # image_wait_to_slice_without_suffix has "color" inside, have to get rid of it
        # delete_object = 'colors'
        # # if delete_object in png_name_list:
        # #     delete_index = png_name_list.index(delete_object)
        # #     testprint(delete_index, end_cus=" : ")
        # #     testprint(png_name_list[delete_index])
        # #     png_name_list.pop(delete_index)
        # #     # del png_name_list[delete_index]
        # #     testprint(delete_index, end_cus=" : ")
        # #     testprint(png_name_list[delete_index])
        # delete_count = 0
        # for items in png_name_list:
        #     if items == delete_object:
        #         delete_index = png_name_list.index(items)
        #         testprint(delete_index, end_cus=" : ")
        #         testprint(png_name_list[delete_index])
        #         png_name_list.pop(delete_index)
        #         # del png_name_list[delete_index]
        #         testprint(delete_index, end_cus=" : ")
        #         testprint(png_name_list[delete_index])
        #         delete_count += 1
        # testprint("There are ", end_cus='')
        # testprint(delete_count, end_cus=" has been deleted.\n")
        #
        # ###Step4 begin========================================:
        # # png_coordinate, png_name_list
        # Type1 = [(280, 90, 610, 420), (680, 90, 1010, 420)]
        # Type2 = [(70, 100, 400, 430), (470, 100, 800, 430), (870, 90, 1200, 420)]
        # Type3 = [(260, 90, 590, 420), (665, 90, 995, 420)]
        # Type4 = [(70, 95, 400, 425), (470, 100, 800, 430), (870, 90, 1200, 420)]
        #
        # def which_type_xml_file_is(xml_file_name, xml_file_path):
        #     type_list = {}
        #     type_list['2017.9.21'] = 1
        #     type_list['2017.9.22'] = 1
        #     type_list['2017.9.25'] = 1
        #     type_list['2017.9.26'] = 1
        #     type_list['2017.9.27'] = 1
        #     type_list['2017.9.28'] = 1
        #     type_list['2017.9.29'] = 1
        #     type_list['2017.9.30'] = 1
        #     type_list['2017.10.9'] = 1
        #     type_list['2017.10.10'] = 1
        #     type_list['2017.10.11'] = 2
        #     type_list['2017.10.12'] = 2
        #     type_list['2017.10.13'] = 2
        #     type_list['2017.10.16'] = 2
        #     type_list['2017.10.17'] = 2
        #     type_list['2017.10.18'] = 2
        #     type_list['2017.10.19'] = 2
        #     type_list['2017.10.20'] = 2
        #     type_list['2017.10.23'] = 2
        #     type_list['2017.10.24'] = 4
        #     type_list['2017.10.25'] = 4
        #     type_list['2017.10.26'] = 4
        #     type_list['2017.10.27'] = 4
        #     type_list['2017.10.30'] = 4
        #     type_list['2017.10.31'] = 4
        #     type_list['2017.11.01'] = 4
        #     type_list['2017.11.02'] = 4
        #     type_list['2017.11.03'] = 3
        #     type_list['2017.11.06'] = 3
        #     type_list['2017.11.07'] = 3
        #     type_list['2017.11.09'] = 3
        #     type_list['2017.11.14'] = 3
        #     type_list['2017.11.16'] = 3
        #     type_list['2017.11.17'] = 1
        #     type_list['2017.11.20'] = 1
        #     type_list['2017.11.24'] = 1
        #     type_list['2017.12.4'] = 1
        #     type_list['2017.12.5'] = 1
        #     type_list['2017.12.6'] = 1
        #     type_list['2017.12.7'] = 1
        #     xml_file_name
        #     folder_search = re.search(r'(.*?(?=\d))((\d|\.)*(?=\/))(.*)', xml_file_path)
        #     testprint(folder_search.group(2), end_cus=" And type is: ")
        #     type_of_this_file = type_list[folder_search.group(2)]
        #     testprint(type_of_this_file)
        #     return folder_search.group(2), type_of_this_file
        #
        # def get_four_quadrant(coodinate):
        #     x_lu, y_lu, x_rd, y_rd = coodinate
        #     quadrant_1 = [(x_lu + x_rd)/2, y_lu, x_rd, (y_lu + y_rd)/2]
        #     quadrant_2 = [x_lu, y_lu, (x_lu + x_rd)/2, (y_lu + y_rd)/2]
        #     quadrant_3 = [x_lu, (y_lu + y_rd)/2, (x_lu + x_rd)/2, y_rd]
        #     quadrant_4 = [(x_lu + x_rd)/2, (y_lu + y_rd)/2, x_rd, y_rd]
        #
        #     quadrant_1_int = []
        #     quadrant_2_int = []
        #     quadrant_3_int = []
        #     quadrant_4_int = []
        #     for ele in quadrant_1:
        #         quadrant_1_int.append(int(ele))
        #     for ele in quadrant_2:
        #         quadrant_2_int.append(int(ele))
        #     for ele in quadrant_3:
        #         quadrant_3_int.append(int(ele))
        #     for ele in quadrant_4:
        #         quadrant_4_int.append(int(ele))
        #     return quadrant_1_int, quadrant_2_int, quadrant_3_int, quadrant_4_int
        #
        # def get_sub_png_lable(png_image_name, sub_image_location, xml_type, corresponding_coordinate):
        #     #sub_image_location should be 0/1/2
        #     Type_noodle_image_coorinates = [[(280, 90, 610, 420), (680, 90, 1010, 420)],
        #                                     [(70, 100, 400, 430), (470, 100, 800, 430), (870, 90, 1200, 420)],
        #                                     [(260, 90, 590, 420), (665, 90, 995, 420)],
        #                                     [(70, 95, 400, 425), (470, 100, 800, 430), (870, 90, 1200, 420)]]
        #     noodle_image_coordinate = Type_noodle_image_coorinates[xml_type - 1][sub_image_location]
        #     coodinates = corresponding_coordinate
        #     quadrant_1_int, quadrant_2_int, quadrant_3_int, quadrant_4_int = get_four_quadrant(noodle_image_coordinate)
        #     quadrant_1_points_count, quadrant_2_points_count, quadrant_3_points_count, quadrant_4_points_count = 0
        #     for [x, y] in coodinates:
        #         counts_value = {}
        #         if (quadrant_1_int[0] <= x) and (quadrant_1_int[2] >= x):
        #             if (quadrant_1_int[1] <= y) and (quadrant_1_int[3] >= y):
        #                 quadrant_1_points_count += 1
        #         elif (quadrant_2_int[0] <= x) and (quadrant_2_int[2] >= x):
        #             if (quadrant_2_int[1] <= y) and (quadrant_2_int[3] >= y):
        #                 quadrant_2_points_count += 1
        #         elif (quadrant_3_int[0] <= x) and (quadrant_3_int[2] >= x):
        #             if (quadrant_3_int[1] <= y) and (quadrant_3_int[3] >= y):
        #                 quadrant_3_points_count += 1
        #         elif (quadrant_4_int[0] <= x) and (quadrant_4_int[2] >= x):
        #             if (quadrant_4_int[1] <= y) and (quadrant_4_int[3] >= y):
        #                 quadrant_4_points_count += 1
        #         counts_value['1'] = quadrant_1_points_count
        #         counts_value['2'] = quadrant_2_points_count
        #         counts_value['3'] = quadrant_3_points_count
        #         counts_value['4'] = quadrant_4_points_count
        #         if counts_value['1'] >= counts_value['2']:
        #             temp1 = [counts_value['1'], 1]
        #             temp2 = [counts_value['2'], 2]
        #         else:
        #             temp1 = [counts_value['2'], 2]
        #             temp2 = [counts_value['1'], 1]
        #         if counts_value['3'] >= counts_value['4']:
        #             temp3 = [counts_value['3'], 3]
        #             temp4 = [counts_value['4'], 4]
        #         else:
        #             temp3 = [counts_value['4'], 4]
        #             temp4 = [counts_value['3'], 3]
        #         if temp1[0] >= temp2[0]:
        #             max_type = temp1
        #         else:
        #             max_type = temp2
        #         if temp3[0] <= temp4[0]:
        #             min_type = temp3
        #         else:
        #             min_type = temp4
        #         testprint(max_type[1], end_cus=' ')
        #         testprint(min_type[1])
        #     return max_type[1], min_type[1]
        #
        # def save_png_with_lable(png_name_without_suffix, xml_type, sub_image_location):
        #     png_name_with_suffix = png_name_without_suffix + '.png'
        #     noodle_image_name_with_suffix_and_path = image_Name_IncludingFolder[png_name_with_suffix][1] + png_name_with_suffix
        #     original_image = Image.open(noodle_image_name_with_suffix_and_path)
        #
        #     noodle_image1, noodle_image2 = split_original_image(original_image)
        #     # noodle_image =  noodle_image_name_with_suffix_and_path
        #     i1, i2, i3, i4, i5, i6 = get_helf_noodle_improved(noodle_image1)
        #     get_sub_png_lable(png_name_without_suffix, sub_image_location, xml_type, corresponding_coordinate)
        #     save_image_to_files_standerd()
        #     # save_image_to_files_standerd_JPEG()
        #     i1, i2, i3, i4, i5, i6 = get_helf_noodle_improved(noodle_image2)
        #
        #
        # can_not_find = 0
        # Q1, Q2, Q3, Q4 = get_four_quadrant(Type1[0])
        # testprint("Quadrant: ", end_cus='\n')
        # testprint(Q2, end_cus='||')
        # testprint(Q1)
        # testprint(Q3, end_cus='||')
        # testprint(Q4)
        # corresponding_coordinate = None
        # corresponding_coordinate_last_time = None
        # for png_image_name in png_name_list:
        #     if png_image_name != delete_object:
        #         try:
        #             corresponding_coordinate = png_coordinate[png_image_name]
        #             #first, judge which type the .xml file blongs to
        #         except:
        #             #the reason that the .xml file coordinate can not be found is these .xml files are empty!!!!!!!
        #             can_not_find += 1
        #             # testprint("Can not find: ", end_cus='')
        #             # testprint(png_image_name, end_cus='.xml\n')
        #             # if (png_image_name + '.xml') in xmlNames:
        #             #     testprint("But can find it in xmlNames list!\nThe origin path in xml_Name_IncludingFolder should be: ", end_cus='')
        #             #     testprint(xml_Name_IncludingFolder[png_image_name + '.xml'])
        #         if corresponding_coordinate != corresponding_coordinate_last_time:
        #             corresponding_coordinate_xml_name = png_image_name + '.xml'
        #             corresponding_coordinate_xml_path = xml_Name_IncludingFolder[corresponding_coordinate_xml_name][1]
        #             corresponding_coordinate_xml_folder, type_of_corresponding_coordinate_xml = which_type_xml_file_is(corresponding_coordinate, corresponding_coordinate_xml_path)
        #             # (type_of_corresponding_coordinate_xml)
        #             # sub_image_location = type_of_corresponding_coordinate_xml
        #             # get_sub_png_lable(png_image_name, sub_image_location, type_of_corresponding_coordinate_xml, corresponding_coordinate)
        #             save_png_with_lable(png_image_name, type_of_corresponding_coordinate_xml)
        #
        #         corresponding_coordinate_last_time = corresponding_coordinate
        #     else:
        #         testprint(delete_object + " still occur")
        #         ##?????why after I delete the "colors", it still occur?
        #     # testprint(corresponding_coordinate)
        # #########check the error and stufff:
        # # testprint("There are ", end_cus='')
        # # testprint(can_not_find, end_cus=" that can not be found!\n")
        # # testprint("imageNames length:")
        # # testprint(len(imageNames))
        # # testprint("xmlNames length:")
        # # testprint(len(xmlNames))
        # # testprint("xml_Name_IncludingFolder length:")
        # # testprint(len(xml_Name_IncludingFolder))
        # # testprint("The length of png_coordinate and png_name_list are :", end_cus='')
        # # testprint(len(png_coordinate), end_cus=' ')
        # # testprint(len(png_name_list))
        # # png_name_list_new = set(png_name_list)
        # # testprint("After set, the length of png_name_list_new is : ", end_cus='')
        # # testprint(len(png_name_list))
        # ###Step4 end==========================================.
        # # xml_file_path
        # # folder_search = re.search(r'(.*(?=\d))(\d*(?=\/)(.*))' ,xml_file_path)
        # # testprint(folder_search.group(2))
        # # testprint(image_Name_IncludingFolder)
        # # get_png_lable(png_name_list[0])
